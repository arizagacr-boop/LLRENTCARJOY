import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="LL Rent a Car Joy", page_icon="🚗", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .stApp { background-color: #05051a; }
    .main { background-color: #05051a; }
    .block-container { padding: 1.5rem 2rem 3rem; }
    section[data-testid="stSidebar"] { background-color: #07071f; border-right: 1px solid #1a1aff33; }
    section[data-testid="stSidebar"] * { color: #a0b4ff !important; }
    section[data-testid="stSidebar"] input { background-color: #0a0a2e !important; border: 1px solid #1a1aff55 !important; color: #ffffff !important; border-radius: 8px !important; }
    h1, h2, h3 { color: #ffffff !important; font-weight: 500 !important; }
    p, span, label, div { color: #c8d4ff; }
    .stFileUploader { border: 1px dashed #1a1aff55 !important; border-radius: 12px !important; background-color: #0a0a2e !important; }
    .stDownloadButton button { background: #1a1aff !important; color: #ffffff !important; border: none !important; border-radius: 10px !important; font-weight: 600 !important; }
    .stButton button { background: #1a1aff !important; color: #ffffff !important; border: none !important; border-radius: 10px !important; font-weight: 600 !important; }
    .stAlert { border-radius: 10px !important; background-color: #0a0a2e !important; border-left: 3px solid #1a1aff !important; }
    .stSelectbox > div > div { background-color: #0a0a2e !important; border: 1px solid #1a1aff33 !important; border-radius: 8px !important; }
    .stMultiSelect > div { background-color: #0a0a2e !important; border: 1px solid #1a1aff33 !important; border-radius: 8px !important; }
    .stDataFrame { border: 1px solid #1a1aff33 !important; border-radius: 12px !important; }
    .streamlit-expanderHeader { color: #7a94ff !important; background-color: #0a0a2e !important; border: 1px solid #1a1aff33 !important; border-radius: 10px !important; }
    hr { border-color: #1a1aff22 !important; }
    .kpi-card { background: #0d0d35; border: 1px solid #1a1aff33; border-top: 2px solid #1a1aff; border-radius: 14px; padding: 1.2rem 1.4rem; margin-bottom: 0.5rem; }
    .kpi-card.green { border-top-color: #00cc88; }
    .kpi-card.red { border-top-color: #ff3333; }
    .kpi-card.amber { border-top-color: #ffaa00; }
    .kpi-label { font-size: 12px; color: #7a94ff; margin: 0 0 6px; letter-spacing: 0.5px; text-transform: uppercase; }
    .kpi-value { font-size: 1.8rem; font-weight: 600; color: #ffffff; font-family: 'DM Mono', monospace; margin: 0; line-height: 1.1; }
    .kpi-delta { font-size: 12px; margin: 6px 0 0; }
    .kpi-delta.pos { color: #00cc88; }
    .kpi-delta.neg { color: #ff4444; }
    .kpi-delta.neu { color: #7a94ff; }
    .section-title { font-size: 13px; font-weight: 500; color: #7a94ff; letter-spacing: 1px; text-transform: uppercase; margin: 0 0 1rem; padding-bottom: 8px; border-bottom: 1px solid #1a1aff22; }
</style>
""", unsafe_allow_html=True)

MONTHS_ES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
BG = "#05051a"
CI = "#1a56ff"
CE = "#ff3333"
CN = "#00cc88"
CA = "#aa44ff"
TC = "#a0b4ff"
CAT_COLORS = ["#1a56ff","#ff3333","#00cc88","#ffaa00","#aa44ff","#ff6699","#00ccff","#ff9933","#66ff99","#cc44ff"]

plt.rcParams.update({
    "figure.facecolor": BG, "axes.facecolor": BG,
    "axes.edgecolor": "#1a1aff33", "axes.labelcolor": TC,
    "xtick.color": TC, "ytick.color": TC,
    "text.color": TC, "grid.color": "#1a1aff18",
    "grid.linestyle": "--", "grid.linewidth": 0.5,
})

# ── PARSE PLANILLA ÚNICA ───────────────────────────────────────────────────────
def parse_planilla(f):
    try:
        name = f.name.lower()
        if name.endswith(".csv"):
            raw = pd.read_csv(f, header=None, encoding="utf-8-sig")
        else:
            raw = pd.read_excel(f, header=None)

        # Encontrar fila de encabezados
        header_row = 0
        for i, row in raw.iterrows():
            vals = [str(v).strip().lower() for v in row.values]
            if "fecha" in vals:
                header_row = i
                break

        headers = [str(v).strip().lower() for v in raw.iloc[header_row].values]
        data = raw.iloc[header_row+1:].reset_index(drop=True)

        # Encontrar indices por nombre — buscar primera y segunda ocurrencia
        fecha_idxs = [i for i,h in enumerate(headers) if h == "fecha"]
        monto_idxs = [i for i,h in enumerate(headers) if h in ["monto","amount","importe","total"]]
        conc_idxs  = [i for i,h in enumerate(headers) if h in ["concepto","categoria","categoría","descripcion","category"]]

        egr_fecha = fecha_idxs[0] if len(fecha_idxs) > 0 else None
        egr_monto = monto_idxs[0] if len(monto_idxs) > 0 else None
        egr_conc  = conc_idxs[0]  if len(conc_idxs)  > 0 else None
        ing_fecha = fecha_idxs[1] if len(fecha_idxs) > 1 else None
        ing_monto = monto_idxs[1] if len(monto_idxs) > 1 else None

        def clean_fecha(s):
            # Convertir todo a string primero para unificar timestamps y texto
            s_str = s.apply(lambda x: x.strftime("%d/%m/%Y") if hasattr(x, "strftime") else str(x).strip())
            parsed = pd.to_datetime(s_str, errors="coerce", dayfirst=True)
            mask = parsed.isna()
            if mask.any():
                parsed[mask] = pd.to_datetime(s_str[mask], errors="coerce", format="%d/%m/%Y")
            if parsed.isna().any():
                parsed2 = pd.to_datetime(s_str, errors="coerce", format="%d/%m/%y")
                parsed = parsed.fillna(parsed2)
            return parsed

        def clean_monto(s):
            return pd.to_numeric(
                s.astype(str).str.replace(r"[.$\s]","",regex=True).str.replace(",",".",regex=False),
                errors="coerce"
            )

        # DF egresos
        egr_df = pd.DataFrame()
        if egr_fecha is not None and egr_monto is not None:
            cols = [egr_fecha, egr_monto] + ([egr_conc] if egr_conc is not None else [])
            tmp = data.iloc[:, cols].copy()
            if egr_conc is not None:
                tmp.columns = ["fecha","monto","concepto"]
            else:
                tmp.columns = ["fecha","monto"]
            tmp["fecha"] = clean_fecha(tmp["fecha"])
            tmp["monto"] = clean_monto(tmp["monto"])
            tmp = tmp.dropna(subset=["fecha","monto"])
            if "concepto" not in tmp.columns:
                tmp["concepto"] = "Sin categoría"
            else:
                tmp["concepto"] = tmp["concepto"].fillna("Sin categoría").astype(str).str.strip()
                noise = ["agregar peajes","agregar seguro","mas gastos ?","nan","","none"]
                tmp = tmp[~tmp["concepto"].str.lower().isin(noise)]
            egr_df = tmp

        # DF ingresos
        ing_df = pd.DataFrame()
        if ing_fecha is not None and ing_monto is not None:
            tmp = data.iloc[:, [ing_fecha, ing_monto]].copy()
            tmp.columns = ["fecha","monto"]
            tmp["fecha"] = clean_fecha(tmp["fecha"])
            tmp["monto"] = clean_monto(tmp["monto"])
            ing_df = tmp.dropna(subset=["fecha","monto"])

        return ing_df, egr_df

    except Exception as e:
        st.error(f"Error leyendo planilla: {e}")
        return pd.DataFrame(), pd.DataFrame()

# ── DEMO DATA ──────────────────────────────────────────────────────────────────
def get_demo():
    categorias = ["Nafta","Mecánico","Seguro","Peajes","GPS","Lavado","Patente"]
    ing_rows, egr_rows = [], []
    import random; random.seed(42)
    ing_totals = {1:380000,2:420000,3:390000,4:510000,5:475000,6:560000,7:640000,8:620000,9:490000,10:530000,11:480000,12:610000}
    egr_totals = {1:180000,2:195000,3:210000,4:230000,5:220000,6:260000,7:290000,8:275000,9:235000,10:250000,11:240000,12:280000}
    for m,t in ing_totals.items():
        n=random.randint(6,12)
        for i in range(n): ing_rows.append({"fecha":pd.Timestamp(f"2024-{m:02d}-{min(i*2+1,28):02d}"),"monto":round(t/n)})
    for m,t in egr_totals.items():
        n=random.randint(4,8)
        for i in range(n):
            egr_rows.append({"fecha":pd.Timestamp(f"2024-{m:02d}-{min(i*3+1,28):02d}"),
                             "monto":round(t/n),"concepto":random.choice(categorias)})
    return pd.DataFrame(ing_rows), pd.DataFrame(egr_rows)

def agg_monthly(df):
    if df.empty: return {}
    d = df.copy(); d["mes"] = d["fecha"].dt.month
    return d.groupby("mes")["monto"].sum().to_dict()

def fmt(n): return f"${n:,.0f}".replace(",",".")

def fig_to_img(fig):
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor=BG)
    buf.seek(0); plt.close(fig); return buf

# ── CHARTS ─────────────────────────────────────────────────────────────────────
def chart_main(months, ing, egr, net):
    lbls = [MONTHS_ES[m] for m in months]
    x = np.arange(len(lbls)); w = 0.35
    fig, ax = plt.subplots(figsize=(10,4))
    ax.bar(x-w/2, ing, w, color=CI, alpha=0.9, label="Ingresos")
    ax.bar(x+w/2, egr, w, color=CE, alpha=0.9, label="Egresos")
    ax2 = ax.twinx()
    ax2.plot(x, net, color=CN, linewidth=2.5, marker="o", markersize=6, label="Ganancia neta")
    ax2.set_ylabel("Ganancia neta", color=CN, fontsize=10)
    ax2.tick_params(colors=CN)
    for spine in ax2.spines.values(): spine.set_edgecolor("#1a1aff22")
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_: f"${v/1000:.0f}k"))
    ax.set_xticks(x); ax.set_xticklabels(lbls, fontsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_: f"${v/1000:.0f}k"))
    ax.grid(axis="y"); ax.set_axisbelow(True)
    for spine in ax.spines.values(): spine.set_edgecolor("#1a1aff22")
    h1 = [mpatches.Patch(color=CI,label="Ingresos"), mpatches.Patch(color=CE,label="Egresos")]
    h2 = [plt.Line2D([0],[0],color=CN,linewidth=2,marker="o",markersize=5,label="Ganancia neta")]
    ax.legend(handles=h1+h2, loc="upper left", framealpha=0.1, labelcolor=TC, fontsize=10)
    fig.tight_layout(); return fig_to_img(fig)

def chart_acum(months, net):
    lbls = [MONTHS_ES[m] for m in months]
    acum = list(np.cumsum(net))
    fig, ax = plt.subplots(figsize=(6,3.5))
    ax.fill_between(range(len(lbls)), acum, alpha=0.15, color=CA)
    ax.plot(range(len(lbls)), acum, color=CA, linewidth=2.5, marker="o", markersize=6)
    ax.set_xticks(range(len(lbls))); ax.set_xticklabels(lbls, fontsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_: f"${v/1000:.0f}k"))
    ax.grid(axis="y"); ax.set_axisbelow(True)
    for spine in ax.spines.values(): spine.set_edgecolor("#1a1aff22")
    fig.tight_layout(); return fig_to_img(fig)

def chart_margen(months, ing, net):
    lbls = [MONTHS_ES[m] for m in months]
    mgn = [round(n/i*100,1) if i else 0 for n,i in zip(net,ing)]
    colors = [CN if m>=30 else "#ffaa00" if m>0 else CE for m in mgn]
    fig, ax = plt.subplots(figsize=(6,3.5))
    bars = ax.bar(range(len(lbls)), mgn, color=colors, alpha=0.9)
    for bar,val in zip(bars,mgn):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5, f"{val}%",
                ha="center", va="bottom", fontsize=9, color=TC)
    ax.axhline(0, color="#1a1aff33", linewidth=1)
    ax.set_xticks(range(len(lbls))); ax.set_xticklabels(lbls, fontsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_: f"{v:.0f}%"))
    ax.grid(axis="y"); ax.set_axisbelow(True)
    for spine in ax.spines.values(): spine.set_edgecolor("#1a1aff22")
    fig.tight_layout(); return fig_to_img(fig)

def chart_categorias(egr_df):
    if egr_df.empty or "concepto" not in egr_df.columns: return None
    cat = egr_df.groupby("concepto")["monto"].sum().sort_values(ascending=False)
    if cat.empty: return None
    colors = CAT_COLORS[:len(cat)]
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
    # Barras horizontales
    bars = ax1.barh(cat.index[::-1], cat.values[::-1], color=colors[::-1], alpha=0.9)
    for bar, val in zip(bars, cat.values[::-1]):
        ax1.text(bar.get_width()+max(cat.values)*0.01, bar.get_y()+bar.get_height()/2,
                 fmt(val), va="center", fontsize=9, color=TC)
    ax1.set_xlabel("Monto ($)", color=TC, fontsize=10)
    ax1.xaxis.set_major_formatter(plt.FuncFormatter(lambda v,_: f"${v/1000:.0f}k"))
    ax1.grid(axis="x"); ax1.set_axisbelow(True)
    for spine in ax1.spines.values(): spine.set_edgecolor("#1a1aff22")
    # Donut
    wedges, texts, autotexts = ax2.pie(
        cat.values, labels=None, colors=colors,
        autopct=lambda p: f"{p:.1f}%" if p > 4 else "",
        pctdistance=0.75, startangle=90,
        wedgeprops=dict(width=0.55, edgecolor=BG, linewidth=2)
    )
    for at in autotexts: at.set_color("#ffffff"); at.set_fontsize(9)
    ax2.legend(cat.index, loc="center left", bbox_to_anchor=(1,0.5),
               framealpha=0.1, labelcolor=TC, fontsize=9)
    fig.tight_layout(); return fig_to_img(fig)

def build_excel(months, ing, egr, net, year):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"
    thin = Side(style="thin", color="1a1aff")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i,h in enumerate(["Mes","Ingresos ($)","Egresos ($)","Ganancia Neta ($)","Margen (%)"],1):
        c = ws.cell(row=1,column=i,value=h)
        c.font=Font(bold=True,color="FFFFFF",name="Calibri")
        c.fill=PatternFill("solid",start_color="1a1aff")
        c.alignment=Alignment(horizontal="center"); c.border=brd
    for i,m in enumerate(months):
        r=i+2; iv=ing[i]; ev=egr[i]; nv=net[i]; mg=round(nv/iv*100,1) if iv else 0
        for j,v in enumerate([MONTHS_ES[m],iv,ev,nv,mg],1):
            c=ws.cell(row=r,column=j,value=v); c.font=Font(name="Calibri"); c.border=brd
            if j==1: c.alignment=Alignment(horizontal="center")
            elif j in [2,3,4]: c.number_format="#,##0"; c.alignment=Alignment(horizontal="right")
            elif j==5: c.number_format='0.0"%"'; c.alignment=Alignment(horizontal="center")
            if j==4: c.font=Font(name="Calibri",color="1A6B1A" if nv>=0 else "A32D2D")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=max(len(str(c.value or "")) for c in col)+4
    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── HEADER ─────────────────────────────────────────────────────────────────────
CAR_IMG = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAE4AiYDASIAAhEBAxEB/8QAHAAAAQQDAQAAAAAAAAAAAAAABQMEBgcAAggB/8QAXBAAAgEDAgMFBAUHBAwMBAYDAQIDAAQRBSEGEjEHE0FRYRQicYEykaGxwRUjQlJystEIJDNiFhclQ1NzdIKSosLhJjQ1RGNkk5Sz0vDxNjdGVEVVVmWDhOIno//EABoBAAMBAQEBAAAAAAAAAAAAAAABAgMEBQb/xAAoEQACAgICAgICAgIDAAAAAAAAAQIRAxIhMRNBBFEiMgVhFCMVQoH/2gAMAwEAAhEDEQA/AK2utgSKsXs7PNp1t/kw/equ7sYHyNWH2cHGmW7EZxa9PnUDSsU490w3OmSMFyV99dvEdfrFUNq0BjllQAjYkV03cBLq2kUqSMZqiOOtN9j1CTCkITlSfI1pfBnNewrpUvtGmRTAkh0B+uinGTcvZrZY2zdAH/WqO8ISs+iCNusbsmPTO32YqRcXjPZpY/5WD+9UtBfBV7dTRfs7H86m/wAbF+41CW2ZqMdnf/GLg+UsWf8AQak1SBFhNtHTjtd/5Uh8/Zl+80iy5j+VOe2AY1aDH/2q/eaILkMi/EqziEZ0ybf9Bvuq0eF/+RU+I+4VVuuAfk6YdPcbr8DVpcMjGlLk+X7oqpvkWLokGkDGl65/kZ+41Q+unMkv+LNX1pXvaVrZ8rI/caoPWhmSX9g/dUroJ9ks7LFBsDt/ziX/AGas/hQY4ksB/XP7pqq+yWGNbQyKCGMkoJz+zVq8KDHEtgcbd4f3TSNEV12lNa/2SXnJcWTSiZ8gahJC4OfFT7ufhtTfg1c3Kt/X/wAMJPBvEbU/7SluX4jvBzakyd+/KO6ilQb+Hj8qacGBhcLzc+7bc0Ij8G/RFaNcIzXZYXD+2s2P+OT76hXa1/8AHWp/438BU44fGdasf8ev31Bu1on+z3VPSX8BU9A3yadiP/Fbz/K3/wDBiq2tCH92rX/P/caqj7DufuLsEjl9pfbHj3MVW7of/K9r/n/uNUlvorXjJR7df5/wz/vVIOyEf8FIfLlH40B4z2vb/O/51/3qO9j7B+FIwM7AD76bModkw1X/AOCOIv8AJ0++qN08f8MNO/bb9x6vTVxjgfiT0tl/eqjdM34w03P+EP7j0vRo+y8h/wAQT1jB+yoX2/DF9Y4//Lo/xqaJk2Cf4sfdUM/lAf8AHdPP/wC2xn7TQEgb2QuRwnEM9Hk/8V6lOvYbs54nPk1rj/tRUV7IlzwnH+3J/wCK9SjX+8HZ5xKqoWXNsWOen50UCKi4KbHE9x8V/eqybh2SJmRuUgHeq14L34pnH9UH/Xqy78D2SQkbcpzindKyGU1ONFTUZ5e6eaVg680y8yqTtnAIO3nWWkRRF1BdRiVInHclYjynbw9R4ijFx7DpmmXLTBJxcBggVMO3U4by+IoTwqZUjFjFZXF1HcNhoXGVK7gnl8W2G/ltXm+be39G0UJWJted0zLG0m6gZBPvb7jp1zv5U7tjHJbIyRiK5CgJ3ZJEhA6kHbm+ysljs0cIbMqsbFC7oPcIP6PlkbemKbMbNnuJbactyklYmAyR44PwrRRtWXskS3gWxe/uDeXOoQpJb7iORSCQdyNhvtUwfVIv7Hb6C0e0t1DgtH3mTIyggsg68h+Oc+AqsLTW5LWweNOaJHZdlx7oA+s0TjzJpdzd98sbIAIFkwjFjvnyI65rlUJeS2UENImubvULvTFMdzaPH3sbxBcgZwd8gZ+P108t9Mm0bh+S5imu2fvsW8NxbAIQdm5jnqdiCp+YoDYanB3bQyJHb3EhHOYBysSPo7joPGjE2qnUIY9Oula6kuGImR5CJNujR42OMfHetf8AtQMN6hNZWoiM7W0jmEANKvIGPiHbHKGOPHrVZ61qUI1Z5oYmWIAB4D7hTBPipIO/iPCkNRW6OpT2sN3cMiyFeSTYkKehHniiH9jyyWpuobi3KRyqCjf0pBPgPHFVSj2QkJ6Rc6jdK6whBDI2XL5JU591l8cg4Py3qTcMWDWeow2N5fQoLlu9cFfzRJ8Oux9cddqixhkbUk0mwtZLgtP3kQhTLuc9M+A+7erSHCel6TwzputJazm5kmkjMVxKQ68smOUjoCB9dY58usKTK1bdm89lZRkpaNbt3Unfwsmc5UEEE9N85+qhF1PFc3U5ntxdLyFliDAnJ6cvljHTIp/qst3pd3+ciSO1mSQd2yYVXYfSz0HhtmowkigI6qtuyv702T+cDDA5T0yD95rkxpzjyXaRMpo0TTTaWKi3trWAYknlC95kbgDOSR8Bvmo9epq1zCrIsbREnmaIcxKqSDnx3GMZ8MUpHJILKS0mmKRK2eflLYYHAB+r76K6Dr9rCkxvZ5GmRmAjKDlaR/oEDyA3ojFwToE1JgmXSJ51jt2tJQ8C83NI4/OKRt7w2z6Vs+lGPTrBZJcmUkLEoPuMOpY/P7KI6U696LlUIiRvdjaTGPh6elEJLm0u4NQt7YGY9Y02Hchhk5PxJrSEpylT6B8Atli0+8hZogJ+VGcF1I6hugJ2wPSmFuYJnjUusoU5IJx18MmlpdQ7m3jhuIo3jjcsC2C+eUgj4dDimdpcpCDJHEWU4UbDwG3/ALVrFVGgsLWEXtXLae3yQqm4Ehwny6fbWyX9vbCSMXHtEmQsL5JKNuSTjqNhTCTT+6tfamlKCVW5CD0xufnvSNpaPZWytDClxPkfSbIwT4D5mlx2xBq3l1Rre7kvBbK0gBfvnAk8wF32GK1cXC20qpcNIofmEasdh4426jNb3M8mn20dxqloJHccy78x5OhXfYffTG1vpJyYni5IBy4xHlsnPQ1GluwsM6dZNB7PeFFKKeU91HzOT5kZ8KkljBZXM1pzWqGRQHdui8zYIOCc5P4VGO9kXS3htnJW5lEYkxy4A8Pn51tZjUVDcofEnK3NI59wY236nPgPCtMfE7synG0TrWJzbWUnKxR+YA8ueb5Yz6VFtYMUd5aPLcyyuIW5mcYzkj7ulPOE7jUu4KzSW56ugfdseO5PXag3F16sl0DcBRC6lUkAAZd84z8q7P8AITnt7OR4nE0OoQuiQK2ZImOCBkMD50hol2kaXXMwRfaiMnxyBTQ6rp8cayrZbxNyiQNlmHrSbR2d/b3DvFdWySS8+Ahww26Hw6VeLNJ2yJRSQR1fUCkZe3tGnDKVZxj3R51YXDFuH0m3d1IZo8nJ8MVVlhZmASwrbXM0PKTG+TgZ86tTQdR0+HToI3mQOYwpVGzg46Vp8d27kZSCc1tCoT3RTWO2jLzKEQbA0RuFR0RlBPzoTczQoDIZhAn9HzFc758a65T1VmcUA7+wkk1iORVIRYkzygHmOTt8aW1eyitbV+55I3WNiGx4nxxRCzlj9pLGdJHCEqQcLgennSGpatD7Q2bZ290fRIPXwrzYZ4Sm7OyWNpIoTUDdG/uQJs8kpX3QQPj0rKsl9A013mup9Md3mmY87TcoI+usrdzgMry9O5+qp92dM35LtOU4Jthn4c1QK83zmp5wASujWhUDPs3j8a1OxEwhUM78hI3x0qB9rGjd7p/taR4K5U/A1NYJZIwSzc7E+VL61p6ahpc1u4yXQgfVVmcuigeDpO7a8t28w4z9R+6pZxf/APLKx2/52B+/UQhhl07iqS2kUjm50IPn1FS/i3/5YWPiPbF/26DNPgq9tnYUa7OP+MXXrJF+61Bm3LUZ7OgfaJwP8JF+61J9FRLEI/N/KnHa/n8sW+D/AM1X7zSB+j8qcdsI/uvbH/qi/eaIdjyfqVVrm+nzePuMN/gatLhkE6XGT4kD7BVX65/ydNsNkf8AdNWrwzy/kpB6j7hRLsjF0HNJGNL10f8AUz9xqhdX/ppP2Pwq/NIH9zteB6exN9xqg9Y/ppP2KSHPslfZOM2H/wDLL/sVaPC23EVhsf6Q/umqv7KDjTxt/fZPuSrQ4VyeI9P8i5/dNI0RWfaWIP7LNQDR6YW799mt54m69SwOD8aT4H7sToFEGOYH807OOjeLU57S5JRxTfqpu2/nDjEGoxEdf1WIIPpSHBQY3Sd4J+bmH9Kylujfq7VZmix+HD/dix/x6fvCoR2tjHH+qf438BU54eQjWbED/Dp+8Kg/a7n+z7VP8Z+AqWwZp2HriC8/yt//AAYqtvRARrdqP2/3GqpuxA5ivF/62/8A4MVW5oufy3af5/7jUin0VlxocajfjH9+f96j/Y+AOFIv2V/GgXGw/ulqIHhM/wC9Ug7IAP7EYT6AffTfBnDsl+rD/gRxL/kq/vCqP0rl/sz0wech/cerx1Yf8COJf8lX76ozS9uM9KP/AEh/8N6V8Gj7LyjGNOjOf72P3RUL/lADN5p5/wD22P7zU2TfTYj/ANGv3Cob2/jNxp+f/wAsT8aAmDex1M8LJ6SSf+K9SbigrF2e8S8zAFjbAEjOfzo8Kj/Y0C3CaADbvZP/ABZKkfFsAk7PuImZsd2Lc7dSe9G3woBrgpjgc/8AC2ck/wB7H79WZfOi28hdC68hyoOM1WPBgxxZLg/3sfP3xVmTAuBGSRzbVTXBDVlO6hL7fqqRT3NzAnMwjDgd3z9M4AG2P/ei1jcmx09YdO1EZE4CyiMhgSFPeZ6gDOMfKo/r88MmszTiSAyxXBVIRgKMHfIz4+lIPqWpT2jWdpGTHndYoskjmLYyBjAJ+NeesN8ejZSo8mvJ2MrTTMXeR1LdCDzbkeWetMxFGt4Oa4IwdiNtsfjW50bXpfoaRqUvP4LayEA/ECiNpwRxjOqmLhfWi/8AkEmMfNa6NG1SFYLjmYuHQIxB6NuBT2QyXXNzDlXwQE8oI6mjNr2bceOwb+xPWAC3X2Nx+FSC27L+Ppbdy3CF/IFACA8qM3rkkfbio8UilIiKypHeQ91bNIQAZyhOfiPlinl1fyXGoRwhA8nKF5ChL7HIOxznFTe07KO0GG1UWvBFws7DkZ5byD7Bz4xUq0Tsv4wskslfhCMXURfnvI7q3Zjt7pwz9DuPlXPOOSDvWx7JlX2LQajqb3k9rL3xXmkWQghjtvgYOwp21nOl6XtbgLa5JJDB3I8yNvCp9J2Pcf3yCQ6VBaSt7rj2yHBXwyA53+Fex9jfaSsQthYaaYg/PzG9jDE4x1zn5UnCcuWidirn9p03WYbiyvGSV4vdeFyNiehq3NflluOyvheOQF7m4nnZ5xnmAV8tsPpbetN73sO42l1HvrTTbG3i5R7smpI5z47gedTDW+zjjZ+BdB0vT7WxN7YtN34a8VQA52wc71eTA5Y+FyVGdcECk1Cye7lFvaW92O45WExYhDjrhicHp51GtNspbuK4dxDcxW2XYMeRs53wBjPyqx7Psm47aHvLrT7VZkUBFjuo2U423JO+1aal2W8ftKxh4egdQCVMV3Cp3GOUgt45FcMMOZcUPZMhdzd2yxXFt7OFjMXIiKS6242PU5z6Z+ukNJaG457aTTnZRIWMytlgMDdtj7vXcevnipLfdlnH6kSxcL3891vk+12xjIPRT+cycb+FL2PZ1x7YIGHDt3GxiYMgZX5SRjAIbOK18M65QWkFp9P0n2GOOScyyXyApFI64iKgDmXIyBjP++oZLo8mm3FzFNfLM7K/LGjj3j4E5zhSPhv5VItJ7PuPoru2mutDnBhQtzPICCRuFO/pSPFHDnEdxLPcS8NagrzFH/M2ze6cjKbZ93qaSjODUXEE02RaB4Rqco5eflt5Cyu3MoPKBjfceJ+Qry1tBPDYxwxhSxCyhj0IXJOfLcUS4q0XVdP5dSjtbmOOSKQSs0BBUgAcrDHj4bULbU4EYA2kbYKEc4wQwGOg33NbJNq6LtBHV5bVLiHkC3EYQSSIikYPQDb4E070e9ENpcWzpGySDmwFywPQE+I+utbUG1t/Yo5o5BIqS3gePB7zm2RT1x0pa7bv7hmt17pwvJImPvB8KwpVUhMEa/362rXLQr3rD6CuXGNh03I862ghupryCCeUG25GPKMbD08qfC2sbeOWOW4jknl3VmJ/0BnGxpxw5BcXur4itYUKr3SyA4AA6/HAz9laReypIVarkUj7lFlijhnuLWI4tx1Iz1J5acG8vVukt5YVhimT3o1jblUbdck4O320n+TbkW8kFrOyxM7DmRgDkEjHXek7eW4gJtJ7mWRuU83OADzH0/hVSxadmO23Qy1B7q4uGmtWMU0KZwUOc+YHgAKG6hHciBFEnfIx3R8qpJHrRe/uby7h9jMKW8cbEFiSC5x026ZxTO7t4bW6dmlLJygKhGcNjcb0mteQ5FdJkjigeBjCQAYyyp9HfI36ncneiM1+RZi2uJp1VwQsiMArfMfxoFNcWltayyweyRmUdHbmIx4jyz038qGyaxJqcSpNMlsIFyj8gGceAA8fWqwbW2Y5YqiSNqkenQCGNpXlePlYd/kZI9BnofOifBEl1qd9axkpGkI5wMc2R039ag2nd3KVYB5WC5keRguCPEZqzOErvTbKyadry1id4sJg5KnrgnxrojCTl/RzSSROdXaeOyUwMFPicb9Ki/EV5DFYqr2ytcFMrztuc7E4p3ccX6emmxTNcq79CoUliPPGKBX91YalfNPLPKsRAZeWJs79V9K6cluNxJx/2MdK0bUL6/kvbu5NrGqgJCznGNvE+GAaP3Fg0oLF5lhUiOXvH5V+PTcGkH4j03Tbpe8gmIaJR3fLvkN13ppI0+ts8k19y2wc8jq3upjGM+Zrx8l45cncmpKkGH0/TkEYa3kuMqSvKmVHTzrKUtNb0axtooItTBdAVkKQM2Ttv02rK6o4W1bMpd8FI3ozmrA7NolfSrTnXIFt/tVArlCWOfOp1wdqNlougWF3f3UdvE8IjVn6Ek9K9A6kyWtZKxCq/KNzsKeW6umQ24xjeo0nE+iS3RP5YtuUrsOfFE4+INJZ1Vb63ZeXAIkFOySrO2DTTp/FFtqkYPcyuCT5HoRSvFef7WFj4/z0fe1SntFsYde4VuUt5oHlt8yx8jjw6/ZUP16bveyHTJv1rxfuak2JrgrlvGjXZ1/xy5/bi/dag2PeNGezoH225/bh/cam+hR6LE/vbH0P3U57ZSfytbD/AKmv3mmx/o2HpTvtowusWoB/5mv3tTg+Qn+hU+ujGmT4/Ub901a/C2+kxn1H3Cqq1050yfAx7rfumrX4XH9xo/NcfuiiXZOIPaUP7m64f+ot9zVQOsf0rfsVf+le7pmucxxmxbH1NXP+qbyMf6hpIc+yXdkw/uev+Ol+5KtHhbbiPTh/0h/dNVb2Tn+YJ6zS/upVpcLHl4l07/Gn900mWive1aNhxPfGSMuGnfHPpfMCM/rg5NM+B+QSoqRJGA3RIDGDs3gd6e9ps1svFWoKtzZxsLl8qmqSRtnPipGAfhTHgwhr3IKt7w3Fx3x6N1atJdGa7LK0DP5asf8AHp+8Kgva9/8AMHVV/wCl/AVN+HmJ1qwz/h0/eFQjtg/+YWrH/pfwFZjE+xAYF5/lLf8AhR1b+ig/l2z9S/7jVUPYiDi8/wAqI/8A+UdXBon/AC5ZfFv3DQWVnxpj8rakP+sSfvVIuyBQOEocf+utAOLY3l1vUY442kc3MmFUZP0jT/gbVV0bQodNeBri+x70MR5u7/bbovUVWrl0YxaRPNVP/Afiby9kX96qX0HTr+64u02SCznlVJMkqhIA5HG/1irS02PWb9+e6uBBZMcSW6KOSUDoGPUj4YqSx6jdWMBS3uYrWNfCOONQPszVrGxyyRs207Q9Wu7CKKGxmJ7sDLLgDYeda8c9l+qcX3FrI10ljHDaLA2U5ySPGhWpcR6k5DNq12V/RUTMC3rsdhQS41K4mdpJ7+4Ync/nW/jVLGiXlJzwd2V2vDWkLYyayJAGZi7ciblmY/vUTveEuGJ9LvNM1DXYfZ7vk71fagre6cjcdN6p2/voM8gdpWI3BOaGBrRyXSNJCfpqwAJo8ZHn9FtadwF2O6NcNdflK0MrDDNLfs+d8+LUcjHZJB75u9FbH6WVb7SKoqGW2XKwKqN1MTbUyfU1W5YQGS0fm+iwJU0/GHks6Isb/slty0tk2jAk7tHEu/1LRCDi3gCIfmLqAAbe6hFc3W9xcO/ePEYmHVoyQDSlxfTzYRZBFH44O5o0K8ldnQeqcc8CXNlNZ/lS6g7xeUyW5kRwPRgNqD8O8SdmHDZlktbvUJ5ZPpyXcss7H/TJqkDcW8Iy+X9SaQN7DK2ORMjopppUS8v0dH/21eBMkoWOPK1NYna1wV0QP/3eudEnth9KHumPiAMU5jkTHvRRSL5jrT1QvOzoQdrfB6nHduB/k9Lx9rPCJGyN/wBjXPcUlv0AQ+YZaXQ2yjIij+qnpYednQI7WuEB1ZR+1Ca2/tt8GdTcQj/+M1zwyWrHJjT6qTkt7Nh/Rr8iaPGHnZ0ava3wS2xu7YfFT/ClE7VOBGx/P7IZON1rmW4trGOB3Yd2APpcxpHgvR77iniCHS9P5gWPNJM492JP12/h41LgkVHJsdWafxvoGuStbcOw2+ozIA0roPchB6cx8/SnkkjK+Sq/6OBn4UN4U0XTuG9Ej07TouSGMZZmPvSt4sx8WNeCWfU5WW1kaK2RsSTgYZz5J/5qVUXYQk1buGELTNznpHGMk/IVo99qEoAitGHrK+PsFJXVxpWi2jT3MsVrEBks595vxJ+FQvVe0oAsmi6HPdL0E91III/iBuxHyFFEuSRNC2sOetsu3hzGkymp496WBvPCGqnvuNe0K7w0V/oGlx5+ilnJOw+bMufqoJa8Tdot3NKrcbwR8hwOTSUA+1qKJ8i+y8ubUM4JgI/zhTG+03Tr/bUdD067ZTnMkascjodxVNS8bdo/Ddzb6jf6zpus6QkoF6jWQhlVM7sCpOcDf5Vd1le2upaba6nZzLNBPEHR1OQykUnFfRpGd9Mjmo8CcJ37SE6VPZyyElmtnKkn4A46+lRq87JrP21r3S9fuhOOi3BwT8WHWj/HGvcS6NeQjRdDs9RgaMtIZrswtzA9F2IO3nUaHavdWXMvEHA+s2irsZbcLdR/6p5vsrKfxsclVFLI7oj/ABRwFxHDqkN3JZS3Nqv9KYiJFO23TfOetMdOtltLsyvaRpbTNyMXkbqBuGU7gHOMeVWXoHabwRqzKlrrkdpcH+9XDGBwfg+KP3dnpWsRc1zb2l8p37xSFf4hlOCa4snwL5izVZU+Clr8adbQLFbwQmeaMNunIqNzemx6ChuqrdNdWiy3QVUAEuIwDk+Xj86n/EvBN5Gq3OiSteGFCpgnx3uPDGAObFVvqtnctGz3hdJGK80Ziy+cHY4GRjFcWT4+RSp8j2iuQhfXdqIFLRuYyCyMWAbIGBzAdNzmh7JFfRJzKEkXLFVH0lPj8aFNczW1u0MMF7IGBBLRbH4b5ofFc6rIQq2Msgj2yQdgavwTguSHkT6CdxpWjKzMZ+QMm3KP0qA3LtZKojbng8uYZP8AuojdpIIGWS3MRbG4Jxnw6+Fa8Pz6TZSOmpQJcvuq4TJVj03JxiunG4tUzCSaYy0i/ZYZGQjlwR5gelHTq3JpCrNNzSxcoWMZJIxjOfOh8U0N5dzR2NtHFDkB1bGSfTFGrqO3k0pGI5pI5Aq5XlOcenj4UTyqPBKwuXKJBwhfyXdlyRqpPLzESOSeXG3wqzbNQbVC+FPLnaq34b1G0sbKU+yhZjGWMrA5x5YG5xT+57SLSytg8mmXTxhc94hUAjzwTmtviZoVrZnPBk+jfjqO4XUfabO2794rOQ8pUNuSBnGKRtdRs47PkSzlIYZYIABnHvHfpvQyDtJttV1Ay2VrIkiR8pjkK++pOfPptSl9ra35SS20+CaRvdUIzYU9QScYBrn+Y4ttp8l4torWgxYX8ETEwA2/OOY++dx0zkfP6qyoqeLLbT782t7ZRW8qxnvFRuh5s7k+O5O1ZXNvP2zfT6IlJhtjW/aGGTsp0jkG/fx4Pl9Kt5owuceRrO0cY7KNJP8A1iP72r1zX0ypbiW+gfHNGcUlYaleNemJmOCGbY+QpW8PNITnwpvpSZ1kD/o5PuoolckjC3EhKmWVUz9FXIH31O+IOeLsZ07kUHlvRgH/AD6jcVthScbfCpbxEoXsesCDj+ej/bqfZT6Kik1WdHINm3xBqS9mlyZbu4YoU/ORDB/Zagz5dmyc0d7OYw19dDG4eH91quRKLHz7p+Fbdut/a2et2guCwzZruBn9I1oo9xv2TT3tvjQ6zahgD/Ml2/zjRFcil0U5q2s6bNp0yJcDJRtsehq4uF3B0hMeIHj/AFRVQa9Z235PmZYFB5Tvy4/RNWxwsuNIj3ONvuFORMGqJFahTputuZOXksXY5+BrnrUNRsml926ib3OgauhLDfTdc3x/c98/Ua511K3SVZXYEtybE/CkinyTvsldGsEwwYd9J0/ZSrU4VPNxNpn+O/2TVR9lERjssxjAE8u3+atWrwm5HE+mZ/w3+yaTBcEL7TfbRxRqGPbintD8uIYJB18N8/I0O4PdlunDrKGyB+ciEZ/S/RFNu1q51BeMtUVLwFfapAFeJWAGem4ob2bz3M95dNI0ZIkVcqgTblJ8KtvghdlucOMDrlgM/wDOEH+sKh3a+Ae0PWMf4UfuipNw1Jya3p5bYe0x7n9oUA7VNJu5ePNX1a81O20/R2m2kYB5H90e6i+JqUrLkNOxQiP24H/7pv8Awk/hVgX/ABPZaHfQuWM96uTFaxDnkfYjp4DfqcCq/wBHWS3t2g0GJ9JsZDzS3c3vXM5wBkA/RGAKf6ZBa6dzR2EOZZm5pZGOZJT5s3U1rGFdmMsqoeSJfatNLNqBFjFPIZGtrd8u5O/vv18egqSaNY29nCp7pIkH0Ywo3phYRpABLKyvJ4eQogO/lUEZAJ2z1NX0c7k2ELvU1hh5y/Ii+Q+ygN9qk0zhACWP0IxuB6n19KfGyg9nuJZz31zEmVP6Mef1R5+tC7YxpdRkDlAbrTsKPEs9SmBkdOXO5ZmpDXbSayMae0hi682QNhRq4uDLCUjGcLkimPERQyQBWD5iBaNxvmmCBGn6UlxF3lzNLG3MeV0xg7+ND9YWC3u+55ncDrMp6H5UbiuUs9Kklhfpkdy/iSTUTWRnmbu5DFKxy0cnRqCR9uUDSL30R6SRn3hT23UchaVw8Q3HOu9NbdEjTvpB3OPpKG2NM76+abKKcJ0A86bY7Hl3qBkPJHtGOnrTCW65UOTg07XSpGiQi4Cuw+jim7cP3jsSbhHC/SAB2pDGDTsx/OF8How6ClYmJG/51f1l6ijWlcJiSRRLdyrHzYYFcD66kNpw/Z2sZiW3xzNhjnmJx45oAhkMjH6D59Gp5G+4GCjenSiesaEtspnRA0PUsmzJ6+ooQ4WEAK/OvgwPWlZI87/lHUE+deCY5znrTLmbzrVnbPWrUgH/ALT/AFq89pxuWwB4mh/OcgDOaF6vfbG3Rs4zznNGw48ju5uLzV9Rt9PsYnmlmkEcEa9ZGOf/AF8Aa6e7KuD7fhHQFtOeOa7kPeXlwB9NvIf1R0AqB/yfuBTp8KcT6nb8uoXSYs0cZNtCw3b0Zx/q7eNXFezmztoo7dA9zKeSBSdi3mf6o6ms27OmMUlRrdtJfXR02FiiKA1xIu/KD0Uf1jv8BSHE2vWXDumfQ5pQvLDAg64G3wApaeex4b0hpriUAc3NI5+lLIfxO3wGBVY673mtalNeXszZaPARVOEXPSiyZyroYalrF5rExvp5u9Zhsp6L6Dypn3snLtCebPi21ErLR4mgjkUTl2UA7DB29ad/kiEMFUTHbJyVrRSRytWAgXAG/NI3jjah2hEtcXSsMnm+l9dS46RbhWeMynzLMKFadpcKT3xjjkZ87YYY8aewag3UIhLEIJlUxSNysCdsYNHP5NepyW2l6rwVeyFptJnLW/MdzbvuuPQdKFXazK8UbW3KwlB3IODQvTLp+Ge13Q9YYhbbUs6fdEHAyfoE/OpkjTFKmXHxjpT3+nlo2YSwE9P0h/7VB5rNI7KQ87seXxO31VbEyMDjOCylTt4ioBr+nCEzoXYdcYG2KSZpkXJHJdD0jVbFI7/T7OcFcZeMFvroPHwhc6VdGThjXL7S35eYRGQyQ9enKelS6ygZbWLDYGM4GaUaANcqGkcEocYG3WjhkRbRH9O7RdZ0OdLPjbTgYCcLf2wJT4keFWDA2gcQWsV+pguo5FHd3MeCceHNUbvdMjmieCctNEww6PupHkRUK9g1fgi+k1LhmR5dPdua501zlWHmg8DUNGsZ32SLisT6bdG1bRXcAZVkYckudsqfvFVXr+uiDUmCcyMhwVHT4VfnDHEGjcXaHiMh4gMSQE4ltX9PEVVPanwPc6deC6Cd9ZSAmOdB442DevrXny+PLm3wbcLogtzrUmoKbV4naNmyrblx6ChQtXS655RiMHI5zj7KI3DS2+mrALju5UkPugbb+ND7jnWNZJJkkZuoXbA9TWSio8I1ST5YatNRjQYiAZgQSqjAI86dHV+8kZORIVZublVulRe3uYzKoKI3Kc5YZA9K0nvWEoSKL3y2QUGMY8vSk0mbwyaqiSRarLNdchu0WHfCny+VSdbbTb3TVjubPvkY8wJODzem/wBlQfTFmjEdyLdG59gxHMc/M1J7LUy6SR3LwxRg4wwAHSufJG3waxaknYE1rRbXS9Thu7eZpoWYNlGHKMHdc/UKJw67ZKFtdKsZblC3v5JCZIGMnqflQq20ddavvZ9MmcwiX3mZ8bf1VxVjcNdn9jp6M11eJIoGWiifLh8efhWzx+SNds81vVgez0HS7q1FxfaeRcMfeEeXJG+GbyJ8qyp22pcO6Pgy6bcXUrAKzyEZHp4VlR/x2V+zRZf6KnlXcg0l2ljl7JdK/wApj+9qdyJktmku05P/APUmkj/rSfe1emV9lOTDmJ+Fe6InNrnKOvdyH7K3mX3j8KW4cTPEC+YilH2U2TEsIW+IjtRviuPHY7YY8L4fe1Nkg/NNtRTjVOXses8f/f8A/mqfY30U7jBcfGj3ZsSb+5I/wkP7jUDb6bDrR3s1wL24/wAZD+69VIiLssfl/Nvj9Sn3bmANZs9/+Zrn6zTYjETeqkCm3bDeNdazbs5AHsigfWacOwm+Cs9cONNmBA3Rv3TVpcMNnR4+u2PuFVTrpzpkw6nlb901Z/DLkaQm/gP3RRIiHCD8EpTSdcIHWxf7jVBXRBhkP9Q/dV/aOqy6XxBzH6Ng5GfgaoC7OIpR/VP3UIp9k37KP+JMP+sSj/VWrL4YDf2T6XyHH578DVY9k5zZMP8ArMn7q1ZOg3cdtxFp0hbJSbJAGcDB/jQ1fQ2ys+1cOONdWDnJF3Jn66S7JbSeV76VI+WBZVLTPsi+75+PyqTcbafoycR6hrWvXPPFNcu9vaRH3pATsDjf5UC1G+vtXhFvKh0vSl2Szt/dZl/rEdPhVxj9mbkkHbziSCG5a04ehXUb5DiS5faGA+nmfhTKG1Y3Pt2rXLahfHcSSfRj9FXwpG1e3trdYrVFjjUYVVGKc2ME99KAmQmd2rSkjnlNvodxPPdzCOIZ+ewqQaXoklyiwwcxlJz3xbAB8s+VOeHNIgRQHGwO48Tt41Io7lIp44IAoYHcn6KD1p8EpfYHjsotL542Y3ExbCqx2GOpNb20xeP3SWYjDP8AgPIU0v5zJczKpJUyEsx6tvW+nsqWoBOMb70FBBCiafelsD80MfXUYln57mMZ5Y+bBYCjJZ5rS9A2/NZCn9LegkWRdpye6S26N0O3hToAuo5IlEg5Rj3ZU3rOIow/sxmUMvdACVOo+NKQKSuIz3bkbxv0Pwr3iiRLOCKYMbeRYB7jfRY+VBKZCuI7rnMdr/TwxZPPHswOaQsowkffXDiRF3XmG4pqDyubucckjZLAHqTWSzFxg7L1AoCja9vWuHIDYVf0RXuhQi91u2tmJ5XcMw9BvQaecQTtnfPgKl3ZpYma4l1NxzISIlx1B6t+FAgrIn593OFYSYA9KcQqFb3s+8MF1G3zryON5rxkQq/vH3W609SBIu7ZThnGSuNhQOxe1mmCEFBOnQsKI6cGBPM3dhT0A3oJe6xpGnjurmURzA7iM5JHrTduONIQAKLlm8+7/wB9Kh0S2VInZowgdDkEsOo9arDiO2j0rWZrGNhJEfeRc7qG/RqRwcd6KSe9jvAT0wgO/wBdBbXTDxFpWr6qJO8vO8LIoPvKBvj5gfZQKgah5xgEY8D6VuIzjLbY8aG2FzzSqTj1A8/Gi2rTw2lsZ3YCMDb1NP0TLugTq96LWLljP51thR3sU4NPFOv+338bPpdjKC4zjv5uoT4DYt9VRDRNP1DiviGHTbMKZ523dvoRJ4s3oB9fSuveAeHLTQtFtdNtIuW3t05VyN2Pix8yTkn/AHVJ0wjqg9ZRR21oXkZUULzM36orzSYzNM2q3CMpmXlt0fYxxevkT1NbalmW4tdPC+7MTLMf6iY2+bMo+uoj2p8axaI9po1uVe+vvpAf3qLxYj16D50UWC+NdW/Kuv2y838yt2IQjfvJAcF8eQOwr15VAHfSRjlhJ9443z0oDrLCKe2MecNgR58qSuy6RoJIy/NkEjqPGmonNJ2HtPvLDFtCbmMDuwz+94Y2pWS5g9oYwXNuA2DhidvnUbsrde4jKheUoBnHWle6Im5P6ufhWmpHJIgwwSHWRyOY8pyPlQzSY2N5cgHctn7a002Rre8CBtpRyH0HnW2kEreXWGBUNgE/E1LQxpr6YvossqZZS2BjFRbtJsZbnhWe4tlYXVmVuYG8eZPeB+z7ammvKpFq4K953gyceHrTK4hjuOaBnVi4IYHxGKpdCTqRZPCerx8Q8IadrUJGLu0jn2OcEjcfWCKZ8X2nPD7QOjIQSN/Wox/J0uTHwdeaBIwMui6jLa4HhGTzp8sGpvxBCZdFuY9+eMHlx4+NZtHTJXyRS1jEUSGFi6AYPOuK9Zue6UAqW5CR6V5M2JFBchAB0NJzzol+pjj5gqEg9A2cZ+qlTMbSHa5LnbHnTa9sBOpYHBPXFZa3KuSWwn9XxNOo7iIcvvgkg4AoIK61zTm4b1mDiPS2MMrSiO6AOEcMcAkeh6mrC4S4h0rizS57SWNGkQmK8tHG6n09PhQfii2W/wBOkiePljkyrA9dx1qv4IdQtXt+INHcpqlqxSeNT7tyq9QfXyptcHRCX2Z2sdnV9ok7anpam50tj4btCT4N5r5H66rG/wBJv44hcXZWGEtuynOa6x4E4l07ivRO9TDFgUngcZKnoQRVYdtPAtzp1rLqGmxtJpbNzTRoMtCf/L91ednxyi7R1wafZSUcPIowSUz1UbmnKWqpEhCssrjJbPh6+VLRXZiCW0cSlBndxkmnNneafZPILgPM5xmKMZD48CfAVkuRqN+z2KFUtBFc3MQkJXbn2G2dxT69gtDpwQmSXIUEMvJufI02TUNIeYzmxeHmx7pPMdj1pxd65b3luttI2bcTAjm25QT41jq3OqNm0ohvgz2trpEtrY91ygLGgy4I6Hwq0bsXiaUs0lpHC6jDom5J86gOmcS2ej6gtnawIkMsZKvGhJ5hvufKpVZ8Vd7bKJgGjYEMcbg16WCKguzkncmR7iDUxJIIl5ZHU+8MNkbelZSXFq2ouQ9vISXOTyHBx61lU0NOgSwyceda9qMeOyLSm2/40n3vUli4et23FxLsSOgrfiXQItZ4StuHpZHjjt5RIsqgEnGdsdPGs94myjKjnWVMuT6U54bTPESjzim/dq1YuyyzW6SSXUppIgwLIYwMjPTOaXg7LbKO+9ss9QmhHI6hCgbHMMHek8q6EoMSCZiI9ae8clR2P2S+J1Af7VHoeEgfpXrFSN8IP41mvcKHVuGodA9sMKwT993/ACZz12x86N0DhI5+Ke83rUo4J1BbLQpZI9Lt7loJFLuCQ5GMnw6jJqYP2RNGxH5bJz/1f/fS+i9ltxp4uMassnfBQw7sqNgd/trP5Fzx/iyYY2mM7TiSOSDubmOaJJSrWyZGG8z8DkUn2kTQ3M9nPGzZa2UFSAPHY9fGpDpfZlIED3l7BNLD/QcgYKg3xkeO5zRbjfgibWnt5bS5gte7iEbAoSGIJ328d65/jLJDJcnwVli5LhFC65n8mzYH6LfcasLh6Tk0qMeO37ope+7J9WntJYk1Gy95WG6P5f76kFvwBqNnZhH1KxOAANnGdvhXoOSZgsckuT3htlk07iIM3/4c5+w1Q17yiJ9wByHr8KvOWzbh7S9VfU7+ytobq1MMcjMQCfhjJ+VUZf30dmzDTwlzNjAuZR7o9VQ/LrTTQ6fBMOy/KQJbzTR2rzyySRJIcO68o3APht41IdZ4tt7KVtK4bjW91E5EkucrGPNjVLada6vq+rAx3JDg5kuTJ0Hicjp8qsLSLBLG39mtdkO7tzAvKfEnx+Vaw56IyOhxbWrG49tu7g3l8es79F9EHgPXrTtgmMHetCGRQMUmHVWzMSFHXzqzlcm+xfTNPuLu+VI1JQEF/LH8anmnQ2ftclpZEKYtiH2NDeCrq2N9AsZVVZ+Ur4n41KLi105b6S7SFA52eUeJ8h51LdFpWMLK7MYmVWwFbEkmM8u3h5mto7tXkVwOSEN0J3J8zTh1hlfPcKqeCjP2+ZrGitEQZiXIOcVCbseoNupkEsjMCpLEnfOKTs2DQ4LBwNio6/H4U9mgt5H5nTAbowPQ+ZoZq+o8O6FcQjU72GznkXmjOS3MPMgdBVOw1C1qB+Tr4j87H3XwZaD26888Z3miz1H0l60S4e1nQ9XE0WnXkbkjDvGDjHqCKNQaLbCQMRJEx6PEcqaq6RLgNbeIyWwAxcRAdMYddqC9qc6WcNjavIW/MiQq3UeQqbrp9vaxCWfZwMZXbNUdxvq0mq8RXV07EoG5IwfBV2FCdhQzMzXNwCRgDw8q3upxGhBIz0pnZzBEdv0vCkb+Y8hJAk5vI7iqIbdiKs0tzzFgGydvCrl4OsPYdFsY5FMcrr3jkdMtvVTcI2Ump65bWaDvFkkUEY3A6mugrbTVjVSsrEjblYUrodWQ8LCk8hnDgNIcPGN6Y61qPselyzLlmGycx8T0qZnhlXkwtxIjFifojG9Q3th0s6PotriXn7245T7uOi0JpjUWiupJ2ZmZmLOxySTmtOcAbgU2AJ3LoPiaxjGDvKpplUx0JQ2xqSdnet2+lcRxteS91Zzo0Ux8gejfIgGoi00SjmyB4ZzScNx7RKIrRTcyscCOJSzE+WBU2GrDWtS2ses3UtlIZLQys0bYxkH0odPe6hr17b2FmjzPkRwwoMsxJ61JeGeyjtA4pljZdGl0mybc3N8e6HL6J9Mn/N+dX92bdkukcGRiYy/lDU3/AKS7kQDlHkg/RH2mi7NFD2BOxngVOErAy6skYv7wjvp13jQeEefD4nrV020fdIMDAxt60glrClv3ToChG4IzmmtrLJbTNZSyFk3eFz1Zf1T6rsPUY9aDQbcVaxa6GG1q+Yrb29pNzY6k8yEKPU4IHriuc7rUbjW9dudd1FcXV1KHKqciJP0UB8lG31nxo12y8UHiTiRtP0+5Y6XoxZWKH3bi425s+YUDl8sk+VV/r+stYlLe35RIwzzkZwPhVRIZb2pyrcPpZTB5EA2+JqvO1LtPTQZPyPpaxz6ime+lO6RZ/R9Tin0Wvm37ONT1kSc1zYxOysT+kcBftJrl661Ca4uJJ5SZJHYszE7knck0pzrojHjttssM9pPEfNzflOf4c2B9XSl4O1PiaJsrqUufl/CqvN236uPnXntMnkBWW8jdYy4Lftj4nikDG7ifHTngQ/bgVOezjtYTUNTNtrMCDvussIxy4HiDXMpupM+FEdG1ae1uVblyhOGIzkA9cU1N+xSxKjtmef2uK2uoCklsWBDo3MSM/ZS11Ar8kqAcy7qfwqIdnGqaVdaFBb6ddxyEskiW6yczRqcdft2O9Td45uUERKCf63WuiPRxSi1IEdlsp07tm4gscMkOr6dDexg+DoxRvsxVu3MRfvEx9OIj6v8A3qnllWy7VOEr0sqmSS4sJADkHnQMoz8Vq7GVQQSfo/Rx61m+zph0V1xCEtNSkjJ5FBBzy52xmhv5VtSeTDE+Dcuwo72kwstxHKq5DpgmolGiSPGiL7/PzE7dPKrqzCSphRwJ42y3dsD1HhW2nJFCSTJ3jDocUj3c+Swl5c745a2CzR5IkI33OAKNRCuqyq0aKqEguM1FLG3FrrM1tn3J372MZ6Hoakdw0ahWnu02bcFhtUa1S8tIdRs7tJudoLhu8xsO7IOdztRq/ZS6Eo4rrhvii24h048trcymLUYhsASRyy48PI/EVeNm0N9ackoWWORMFTuGB/A1Qevazbatp0ul2IZ2nB/owZW+kCCFQH76tjs0k1F+HLcX1jd2rwqI83EZQyY/SAO4z61M0ujbHKVcla9pPAFlwzeXF7Z6XHc6bdA8plbe1bGTg+XiPqqsdOtWtdStWcAhwykcvjjP3E12JqNha6rp89ldqskMy8rqfKqQ4r7KtY0i4S60pJtUtVkDDBBkQHYgr47HqM9K5niSdot20VXcWUj3V7cRRRZUgMeU4HoKDXywwRusvIUkYcoRh1qY6lYXcGlSJcI8cjnndHUqc58c1ELm3/NqjH3AxPMRvWco0yYTk+B3wzqMrXhVO7ZY1BbnbG3jUy0NX1KYwJK8kXJu4XBz6YqDaNcSRRC2hYBxJzcmDlz51aHDQNtaBSW5tiQVwcYGB6dawUn5KN3dCeucO30ltZ+xwd8FQh5MZJOelZUnbUrbS0jkEkRSQe6sj4PxrK66+iLZVdhc9p7xe0RtpzAnO8a04W77UnJ20wfGNas/TrJYrdgg90MR8aQE1o90lqs6NKASUB32OPvrl2Sqzuorea57UE0+W+f8l91C6ox5BncgDx9aWtNQ7UAMLHpGwzuvn86n+r2TyQRIsrrHJIoZQdm3GM0+SylidVIUkLg0+eiStG1LtQC5EOkEnwA/30m+qdqC3ItwmlKzADYfPzq1oLFSgLCht5p6niGFkzkxZ+GKTRSIPJq3apjMlvpEuOnh+NapxB2n9DpGkEDzlNWiLBcbkYrLXR5p3dYIiQf03XAFWotickivF1rtNgfvRp2ig92CyiZhgHfyo1Hcdq0yAR6HokxIDbXDjr8qn1hwzaoOfUpjdP8AqAYTA6DHWjaTrCojh5URRgAVosb9mM8v0QPhfRO0G/5n4jj0jSoskBLd2mkI889BUmHDs8UoK3EbgdSw3ok1zO5whdz6DNKdxccvNcypGPXc1Xisy8kmV3xT2O6XxNqEl/qnEGopK/RU5CqDyUEbCgL/AMnHh52DJxJqRx0zBG1W691axErGrSsP0jsK0GoSE/SC+grTRITkyqx/J20wRcicS3iKc/RtkB+ylbP+T7Y2yoF4puXZDlXa1QsPnmrPa8fwbr61obyQdCaajRLd9kCHYdZEfnOJ7s/CED8aa3XYJbP/AEHFlwCf14M/jVim7kPUn6689pfzNXRP4lb2nYTf2szS2/G8sJIx7lqB9uacTdi+typyntDv8AYH5rp/rVYIupAMDNYZ7hv02HzoqxlbHsO1kqo/tiX2226H/wA1I3HYJqMy4l7Qrxt87xn/AM9WeGl8Wb669xKfE/XT1Aqr+0JfJkDj+fBPjD//AJUMv/5NNpeyiW743mkkAxk22dvL6VXSI2I3Jr0RHO52qaQFM2n8m+wtkKLxzfIv/Rxcv40+g7A4YEUQ9o2uR4/UBH41bYi22OK3SF23GSPM06Qir07F4sq0/aJrs3L0zEp+80i3YXoWctxTqbE9c20e9Wz3OB139OlZ3XmaEg/8KgPYRoIJ/wCEmqZ81hQUm3YLoBYN/ZBquf2U/hVxd1k9c1t3S+VFCoqG07DdGtJ1mtuJNZhlXo8ZRSPmBRM9kluwAk414oIHTFyo/CrK7nfNbrESdx9lA6K2i7JbQDl/sy4qI/ysfwpSTsY4eu4xHf65xDeKDkCa9LDPnVkpER03+VaT3NtaDN1cRx+hO/1UIfRXCdg3Axb3pNYb/wDusPup7F2FdnowDbanJ8dRl/A1KpuJIBlbSB5T4M+wrS2m1fUZMKzKp8FGAKYWR+HsZ7MrKdJX06dmQ5Cy30rg/FS2D86k0drp+hWROg6fbwsfcjAt0jDHy5lAPqT6U+htLSyHeXs6u/XGayKKTUZ1u+RkhVeWFCMcq+fxNDQD+C8mn8OUnriidunLHzOcnypvbRpAmwFbPLmkApJLzHAquO3DiebQ9EtbDTH5dWvZCsLDrEmPff6ifnVgA82MbEmuXuP+IH1XjvVtSuJcrHM1nax5z3UUZIPzYgn4YprsHwhiZLfTLBVjRnCDl5QMlj45qEasNRuZ1lazaPlX3CV2IHxqR3XFFhYxkv8AnGHQZA++oHxb2hXt7KEt7aCNUyAWJYnP1VUmkRFNnmq8RazFot1oa3HJY3OBPEAN8dKhRht1OTzk+tK+0ajql2sKiSaaRsLHGhJJPgAN6l/D/ZVxzrCiSLh69VD9EyQup+0Vg+eTdLVEM5LXH0D8zXv83A2iT6s1bmm/yfOO7oAvZG33/Ti/EsKkumfya+Izj2yY4P6kkcYH18xpasWxQKOOiQ5P7FKwLeSyKkUDAueUZ2+qumrD+TfMCO/FqMf4W+d8/EKi/fUk0fsESxnEo1DSISpGGj093Yf5zyfhVKP2JysjXYtwuvDPDov9TnhjvLoqVVnGUUkZNWTca/okIQflWD3Rg8pya8Tsfs3dWuuJtVYeKQxRIPrKk0SteyPhJGDSy6xcEfr3hUH5KoraMklSMHBt2QbjHW9Nnm0a80yXvriy1WC5bEZHuAkOc48jVg6h2r8KQgrFLcTN4AAL95pzb9mXBMRJOid7tj87cSt9haiEPA3BkZUrwvpDFehktVkP+sDSbRUU0qK71/tX0LWUSG2snM/McKzcx22xhaG2eq67dH+5nC9+7FSMi1YZ/wBLFXhZWdpp4CWNnbWsfgsMSoB9Qp4HY5yxOaFJjcEylodK7TL5kC6THaREY5ri4WPHyAJolbdnPGF4udU4hsrcZ+jDE0xx/nEVbByfE1srADFLZiUEiuYOyXS3ydR13V7wsMFUkEK/Ll3+2jGl9nPBunOkkehwXEydJbsmdvrfNS8AVnu+NLkvgZwWcECd3BbRQKOixIEH2YrcwSMc4+ynRkHgK87w+VKgEoIZUbY9adLE4GCRSYl33FbrMM0UNOiLdpHB8fE2hywRMkF6qkwzFRscdG9K5i1rhbiK0v8A2K+tFhuY2POsh90eRHmK7KEgPliozx7wZp3F2mm3nBiuYwWguUOGiOMfMelY5oNq0XHX2cm2mlajp2pGee0spVYcrc0uEG/Wpg9wRG81vNAJ+YCRI8nm2GNz18KAcT8D6ro97Pb3V2kkkTHMbNgMu+CD6+XhSunXGl2NrErEwXDEZQ5O3j4150ZKM3Jm0kl6H6JLfxA93JlfMZx6Vla2HFENi00cLW0nM2Q0gIAHl8ayt18qNcmLtdDnVONbPQteuYGuXuY8AKibqi7k+vNnr8acdntldzapc6yVPc3rF1d9pMEZCgZ2BO9VbolrLPeqsFj3/OQGmdiAoY4UE9fl61aPBlxbaLYsssDQP3zKTuwkxvkZ6DxryIZEsq2fB21xZOL6JvZIRgE98ny94U8MkTzgxyKwwehznc1X/E3GjvaH2KGe3EZWSOZlICMMkHPQjptUMm4z1E3KvZSlUf3WdcsS53OB02yfrrfJ/ILb/UrJquy7ItWs5rsWMLGSXbmx0X4nwpa3tmm1cNHGXkEWAPLeo32ecL6zfd1qmpd7b8558Od38mb5YwKtS0tLeyjxGPfP0mI3Nd/xPJkjeRUZTyJcIH2ulpHytOeZx4DoKdtJyryKAAKUmnA+NM5JSTkDevRUUczk2YVZ/HArxBaxbuxc+WdqayzSyNiNgT0J8BXgXA3PMfOqoQ+a+5FxGoTywKZySNKSXJJNaYbxr0DaihMzl2rFjUb43rdRW/uhcscU6ASKCsx5UlLdw55UOfWvBcqB1FArHAi5tyaUSJTTZbpMV77co2xSofA7ES9K37tRTE3w8MV4b3ahAEgq16Ag8aFi9bmwKVM3cr3tzIIx4A9TTCwgOXP3V60kaDMjKo9aA3GrsxxbDkH6xO9NGuCW5mYlvEmlqFkj9tgU5VCfVun1Vo2oA/pZqPtcnHWvY5GkOAaKFYcF7zuFA3NPojgb+NBrRe7HM2C1O/aj0FFBbCWUr0EddqA3muWdoD30wLfqruaCXvFl5KStlEsK/rHdv4VSQWTmeeC3j7yeVI182OKA6hxbYwZS1Vrlh49F+uoZI15eSc9w8krH9Y5xTy10i4lwAm58xRqkFju74g1O9JVZO4Q+Ee3203gjeV+Z2JY+JOTRez4bnIBcqoova8PxoAWkFJ16BDLRrSDvl7/6PjUomj/max2Lqm++DjIpvBpdrGBlyTW90qLIlnaMBPIMlj+gni34D1NKygbDC090VMZa3ib85vnnb9Xr0HU/Lyo/Hdqox3ZXPpWQQx29usEaKkaDCj5/edyfia1cpjwqbAVE8beJHyrcKG+iw+qmikeFOI3cD0p2BuI2yNqqzj3gPheXVpb6+0aOSaduYtzFQx8c4NWkZWDe7moX2j3V4fZ42TEIOQ2OpppgQ9uzbgeDRYdSl0G3mebmCpvhcHAySaCjg/htcpHolgPEHuRUpvLqcaTDEHPdnYDyGcn7aYQSYbDHqMZp0hWzXsw4P0K04+j1D8h2klwkR5Ze7/o8dGA6A1eDS+RP11Dezq2k7m5upFCgsEQnqcdalbjlFJ9jtm5kO9al/OkuY1hOeu1KgFe8HpWd4PSkfhWeGTt8aAFufNeqxpm1zBFs88S/tMBTeTWdNiPv3sOfIHNFAFg29bKaAScTaUgz3zv+yhps/F9gG/N287n1wKKYEqzkda2jbIzUMfjJgT3dj/pPWk3FV77Ok0UcKB2YH3c4wB6+tFMCc81bVXMvFGrMMCZV/ZQU3bXdUkyHvJfkcUUBZxO25xSUl1bx/wBJcRL8XFVa93cSnMk8r/Fya0y2cg/XQBZcmsaZHnmvYfk1M5eJdJQkCdnP9VDVfnbp19KIW2mSyWouZZo7eFujyHr8KYEll4ssxtHbyt6nApCTitu6JjtgCTgczVGbuGGBuWK4WcfrKCB9tZMMCJD4Jn5migDj8SX8nSVU8goqIdp3FfEtlwvNdaZcziSJ1ZzG2DyeNEhikb+3iuLV4J15opFKuPNT1pONo3wZfFkUilhqGrcVutst/bC7aTvA11ccpPgcE9fhUt07gqzTToVvLa3ubsbzyKxIJ8lwQPnVTa5pl5w3xrNpjFmWGQNCSPpRndT9W3yq+9Je4ktYTHZz8pQYwm3SuLHig27R6v8AKRjNQzQ9kI4q7OLmd7WfTJtMsYZIsvHNdcnvhmGQD6Y3rKs6+0K01jT7Rrq2zJCXX313AJzWVp4oHjJyOe9BN1CLaJHigMr+yjLYb3huxwRsB4+OOlSDRNZZ7x7K9CSXVt+bWQD6AI3UDxNVwdTsfaAZ5JWDY5jtkb45l9d6nOm/ky3k/KaXkVvcRrGwBAYvgb5HQk183nwqS/JHdCSqw3w9FrXEFt7Pb25KGQhbdwq8/Nv0zkAKevrVzcE8EaPw7bI5tI5brYnmAKIcY90fj40l2baWtnoUV9PAI7m4QEkpylU/RGPDbwqUPMoU77V7PwvgQhHdrlnLly7MXaQk9NseFN5ZCTsfrptNeog2ND7jUVOQGr1FE52EJWRRknemNxcrI/dRnCj6TUMub9myAxpuLsKvUCqoLDiSRxDlTFJvcqPKgUl+T0am0l6xPWgLJE14vnWou1J+kKjhvDj6VayXwjjLk0UJskk2pQ28Zd2GB4eJoLe61JcNjJVM7Co9cX008vM+SPCtO9bxq9RB5b7A2IrYX5PjQDvj61uJfjRqAfS9bzrb21vOgKzNge9W6zHG5ptIA2Lwk/SpaGdnO74HjvQATHqScVjXDvsCQKWoElfVoLVCtqgkn/XY7LQya7muJe9mdnbwJ8KGqxpVXIFOkA+WUgV6ZjTISnpmtu8wN6KAfROXO5ojbvHEmSwHxoB7SVGwrSSZ5Bu5x5ZooGGrrWoYshMu3p0oXdane3ZKBjGnktNSUAyK9EhAAG2aKQHsVqpPNI+56770/gS0iG4JNDi5HVt6Yahr2laapfUNRtrYYP05QD9WaAJZFfQRjCRinC6s4+jgVUWo9q/Dlu5SwF3qbjwgiwv1nagd12qa9cHGn6JbWqk4DXMhZvjhanZFal/pq8x/Sb5U5t9RkY552IPxrmi44n4y1HKvrc0Ck55LOELgfE0tZ6fqV1KHu73UZ2zj8/evg/IVPLDhHTlzrFnp1jJeX13HFFGATlxk7gADPiSQPnSmkaxolurT3es6Wt5OQ0gN2hIHgo36CqH03hO3j5jqdvBBbkBpXm5ioXruWNFJtS4Lt2isdKisri5kPKrW9j3ufn4D1oofZfEGo2l8383v7WUeUcysfqBp0EVeuTXL3E3E2j8NI7pBYXVyq5MImEcnhsOQHHzIqT8BdpHFEnDqXM1lFbxSt/N4pmaRgn6wJ3wfXyqaGX/GV8FxSvMvKN6pSXjviWbpcxxeiRAUzn4i4huGJfVbnB8A2BT1EXuZEUbn7Kj3GsI1CwjhS4tV5X5iJJQv2mqekvdRl/pLy4YePNITXgV3OXYt8TRVCsnut/k2PTbSwgvrJniBM0vejcnflAGdh50JtfyKsoN3qiuo6rbxMxPzIFR0Q0tFGopgWFb8baTZxmK2t7l4R9AFQMefjXknaCjDEOmOT/Xm/ACoMsaYx0rdAg2FAyWvxrqDnMdtbRjwyGJ+8Um/FOsSbiaNP2Yx+NRsN5UoH2xSYBiTXdVk+nfTfI4+6ms19cyD355HP9ZyaY95ivS4xSELlyTvvXhkI3pAyjFJtL64qhjsyk+NSHT9PsbbTI9Rvobi7kl/obeLbPqSPCogJMn6QqS2OhcU6jFDI2pQwW4XEeGJbl+AoTAd3Imnt5SdFtbKEIW52U5BHQZz1+VBfaM2aofBy31j/dUg1vSLHRtGkluLiW5umHKjSNjJPiF+uoeJfcIJ8f40APVl360rzihwkGQRSizUAPg/ka9Ep6gimXeeterJg5oAe96cU6fWJppUQab3kcScsayE4Hn0I60MDjHrRu24mmggWNLGzDquO8MeSaAM1SCWJbdpY4onmj5zFGCOXfoc03u2xclR+iFX6hWkl7cahqCzXD8zsw+GPSmxkLSMxOSzE0AOkYZ3rZ2ypFNlb66VH0TmgCD9qnD630Ola1BGDPDMLSVvNH+iT8GH21Zell7WygtlhIEcYXJbHQU3sNPj1Oyls5AuJdgW6BgQVPyNOWuIEJBcBgSCMdPSsJR1lZ1P5Mp4lj+h2skrDPdr/pVlMVukP0Qx+C1lSZWcz2UGjwB5LbToBPbZJJ94tjIGD+l1+ypt2VcKvxXxxa3t/FMNL0qMStFOgDMR9BWx9dAOL+GtEFvZXelXslkHIHLzq4kUnBZQNwQPtwMCrZ7ArVdN4G1md++9oN4Ypmk2yVUDIz4Y3rz/AIfxqyLZ2XmlUSXahr0CXrQzSpHliIwTjPhTW5vyfosMfGqr4k4kgvtfmgltBPptu4BlLLynHUnPkdvlVZdrWqX9rqcGocP39zaW9xAMRwTELzKSDjB8Tn6q9lukcqVnSUt0zHlDb+QNNpJn5tga5HteM+LFCsnEupIfFWlP40+j7QuNovocRXJ+JQ/hUeQrU6iZnLEkmk27w+BFczR9pfHq7rrkh9GhQ/hT6HtY46QANdWsn7dqPwIqlkQtTodg4ak2Vgc1Qi9r/Gw+kmmOPW3I/wBqncXbPxMo/O6Tp0nw5l/GjyoWjLvKNimlwrytjcKKqVe2/WQAG4atD5kTkfhT+07XbqQB34aiJPgLo/8AlqlND1ZY4t29a9MDeVQJe1wH6fDL7eV0P/LW/wDbajI/+GpR/wD2P/8AGn5ELVk8EBHnWGE1BV7WIz04al/7yP4Vt/bWi/8A01P/AN4H8KN4hqydLEdqVERI2FQAdq65/wDhm4x/lC/wrY9q6L/9NT/95X+FPdD1ZPxAx61sLcjpVfjtcjzvwzcf95X+Fbr2vQj/AOl7r/vC/wAKN0LVk/SEitu6Plmq/wD7cEHhwtcg+tyo/CkLntos7ZAZ+HZlJOFHtSkn0AxRuh6ssfu2G+K8dHHUbVXlj2tT6lIYtO4SnlYDJLXICj4nFErDV+0TiBzDpejaTYuT9KSZpeUeZxgD66d8WJqiXlGzuKY6lqumaYhfUL+2tlG+HkAP1VF+ItLu9PQjizj2eWXr7DpCCMn0Lfoj1NCdIuuCbSUPNwi9w5OVnuJxdN/rED6qS2ZNpBe47RtDkkNtpUN9q8xOFW1gLAn0JozBpnaPq0SSRaTp2g28hyrX0hklxjIPIMffWabxXw4WWGzv00rY4/uc8SjbO7AED66lnD0moySGaw1Wx1eMk80YuQxA93ON8/8AvQ4cFKSfopXjXhrtQtWnGvXmoCyiJ5pbCPkiK+ZZfuNQm30fSxJ3nvzSg5LSMHJ+fUV2bDqE4v4rNAkkLhg6uRgL5E9DjxqNcY8J8KXcXtc2l6aVYmSSRQFHqcgjGCOvrSUV7HZzRDHaKuEjTI6jxpZJETLKgHjsN6d8VfkP8vSpw6tw9iq8vNK/MruDuyHryj1r23tkETTSnDBC6x8pJz4e75nwH14ockhKLk6Qra3PdQ99PLFBHnHNIck/BRux+qjGgardvq6W9jp91dTjBVghZ1Hn3Yxj/O2pLhfQNCvbgz3q3l9KuO8b2gR+7jOE9fQUL7QON0jup+H+FrGXRtFjfkaMKRPcsP0nPUqeoX13zUeVS6KeLXskvG+o2fNLFqeqz2kRC5jaQTyZ6nmjjwoz+14DINVFqfGGsKsulaVcR2VnzEfza3W3Lrn9Ll36etO7bT5tVWW71G69hs7dOZnZMkk7BVGcsxPw8amnYt2TXHEuoNq2sq1ro8Tc+X91p8eOOoXap7KXQN7KOA/b8a5rMcgsFOY42JBuW9fSrcCCaX3VHgqoq4AHgAKL6v7NeX0Om6LbKtrboIYFUdfM/wDrwqccMcPW+mRLLMiyXW2ZD0X0FNIT5IlpnBmo3QDShLVDvlxk/VRhuAR3fual7/8AWh2++pdqGqafpsJe7uYoRjYM25+VRe77Q9NRiILW4lA/SOFB+unQqojOs8L6vp2X7gXMI/ThySB6igyMAvXocH0qdQ9oWnyH8/Z3CDzUhtq11HTNH4nge70S4iW9G7J9Et6EGmBDA9bKxG9ISJJbzPDOjRyIxVlPUGtwwoAcpJtWyv8AXTZWHnXvOB40EjxW9a27ymayb1vz0DTHIfJxXpbFNRJg160gI2oCxxzitJGGdqR561LkmgLFCxojba7q0NotrFqNykCfRQOcChJNYDQKx3LcyztzSyPI36zMSftrwN7ppqJAvWsaYcu3nQFj5X92vA2+KQifmGxpdRQUKK3lSiNvvSPTxrR5R0HWgLHgYHxrdSM9aYCRhSqOxHSgVhW1ZQzOD9BCfn0pANvWtsSLSZiN25VH116kTE7mgaY4t072UJzYz0pwyPG7RyZ5l603jiK+95UXkT2qyWUj89FgNjxFAIecOgCE+r7fAUI1+7jttYuYgOXDByefGMjP40b0eMxhF/V6/GhnElkJdYndlJX3dvlUTLi+QbFqZbPK8fycmsreOyhUYVCPlWVkVRTUXDUDaJPbX0tzqAjYm3lj5YYlcDGQ3NhuvjVpcFC80DsCaXUUMVwnfyHPgoY8u/jsBvUC7K+GrHifjuwtde05BEsEt77KpIjCpggED3SGJ3FXJ2kuj8A6tYzWzNE1uVQRLk4PUYFZ/FhzsTllZQHA+omPgu2jnS2adrq4m97l55ixTOAdzv4etHtM0Dh3V7pk4kg72JU/NRlyFi3zjb1P10J4d0NdEsneWdpVdzLCrjAjUjqQRscVBeJe0K4t9TePS44WhTI7yTJMhHUjBG1dfCXJmrfRb8nZR2fXmRbJLHv/AHq6YY+ukJOw3hdwe5ub5c+JYP8AfVM2varrUJzNZQOvxZf40UtO2vUrfdbEjH6lwR+FTcStZFiv2A6VkmPV7jfoGiX8KaTfyflIPc64q/GNvwNAbD+UJdxDEtjKf/5FP3ijdt/KNtFIE+kyN54UZ++i4i1kN5ewDUwCIdYt2x0yXFRfiLsvutFvFtLjW7TvCvNszED7KsQfyiNAMLr+T7qCZlIRu7BCtg4JGd98fVUCt+I7PXtViT+yd7m5lKove6ays+PUSYyfh41nlbUf9atl44Jy/IxODNKilFuWN8EiUtOshVCxG4wR4fhT3XuzvQfZIY9M1REumz3g5WIG22CB1zUn4ej5AEvdI1SIsxJ/mvMMDochumaxeKbPgziy2XiC40aQQsHaE282SPRg5GR8K4MEvl29kdeSODimR6Dsh4rWOOOK6dRyZAdQp5c48R99bN2ScZLsLlD80q2I+2fs4kDzHiHkllPNJmB+vgPgBt/6Nbr2ydnB/wDqGI/GBv4V6Sr2cbTvgqQdk3Gn/wBwn+pWf2qeNlI5bpP9WrgXtg7OW/8AqG1+cTj8K3/tudnGf/iKz/0X/wDLT/AVMp4dlnG4P9PF8cLXp7L+OCc99Afiq1cg7WOzk/S4jsfqf/y1n9tfs2//AFJYj/S/hT/EnVlN/wBrDjjwMB+CrWh7MeOR+jCceiGroXtV7NCd+JrEf6X8KUj7TuzWQ4XiewJPTOf4UfiFMpCXs240C78q+qxpvSGidnDnWxFq7yo77B3G7n9RfAV0bo2s8M8QZXRtWt7pskAIepHUDPjSGt6NHcI0cqZBGD5/7qdRB2ivuH+HLOHUU0+4tvYrJD9GNd/iT6+dS7i/Stdk0iCz4YVbWyj5ueOCQCR84xv4jr9dOdN1WPTQLDiW3afTwuF1FRzSRekniwx+kN/OplpumpBAt1YTRXllKoZGQhlIPiDWm6qiGmzna84Z1iGXupLaQufe94dfTPjTeThfVu75/ZHORuBhcDHnXR2odyEJdR6gjpUK16QexzNpaWrXSbiO4LNGV+VWmRRWOncI687juYo4sgHmdsjGOtTbQOGL/TLNrzV+J7PTk/SBIXZQCcZ6nb6qgWrcVcdi4ksg62PKrLyCFcKD9HHmPHIpnHp5vpfatdvmu5y2QZmyMlcZwTgEgnoKHyVHgumKwNxIzflC3exnTEUhkGCCpKnY7bdfWqe7btevbK30rhUXEXs8VsZJ4oHJZnZzhXPlsDj1r3WuMrLg3S3i0VoZdTuS3dKPf7leUAMfDbfAqp210x3F3qV9zXerSvzKZveAY7mRiepHgOlYt0y0rD9pzWt3aWk0aPqd047i3XH5ofrv8skL9dTTULQWnD50e1t/7o3NwDJOXBd0AJIJ8Oo28c1TXD19rEPEKarDDPfXBdufljL82eu+NvjVmW78f61cSS6fodvp8Um3PcjLb4yRnO+w8K5cylOSS6OzC4Y4tvsf8S2XEHC/C1nFNbwQQysXj5V5pTgZJds7DyqB6zcXep8QRNocDXd3PbqG7gc3K3Q5x4/7qsmx7K9a1+5SXiniK+1DPWCMkJ8//arS4d0ThLgewWO0tYklx7yxqGdj6mqx4HCVsjJnUopJEM7NeyCWTuNW4yeJ0t1BjtEGIwQOrebVLuLddjlRdM0hRFbRe6XQY5/DlGP0aQ4i4lvdVja2jUW1n0ESn6X7R8fh0rTgbS49R1tO+BFvCO8bfGcdB9ddBzko4B0dbW19ulQNNKPcB3KDx+BpDi3jVLFntNNxJcDYyZyqfDzNNOO+Io7ZpLHSm5XkH85dDsfID18z41XEsuW3GM+VArHt5qFxeTtLcyvK5OeZj/6xSBnxneoNxXx5pmjc0Nvm9ulJHIpwin1b8KgF/wBo3E1y+YbtbRM5CQoAPtyalySLjFsvhJgd8Zpzb3M9vOs0EjxuDkMpwR865907tH4ntZxJLfC6XbKTRqR8tsj66tvgjjDTuJbYiJfZrxVzJAxyPXlPiKFJMbi0WBqOpPrLRz3CoLuNArOox3mPE+tMTttTNZArggkEGnMsisofxqjM258CtTLg70kZDiknYlhmgBz35z6VvHcHx3FMhmt8kDagKCSOCM1sX2pgj4rcuxHpQIclwPGte9AO4pAb71hGfGgByJFIrcMCOtM1XelFBoA2kOWrAMp862RObalViwmw8aKA1gZlpdZTjetY4xjfbelVh2ziqUQMVifGvQCTvSqR4FbhN81VFCSjenUS4wQK9ijHlTq2iLyqnmcZ9PH7KTQG/II4IlbqcufnsPs++qx487YNK4fu30/S4PyldxsVkbm5Y0PlnxNSHtU1uez0p7exmWK4uTu+MlE6ZA8/D66591GOwu1dZAkjrtzOQrfLA8PtrCc2uDSKLW4I7boNT1NLLWbKKzSUhUmjYlQT+tmr10iQOFIwVYY65BzXBs8JtLkd2wZGOzDxHlXVX8m/iZ9b4XFldvzXOmnkZid2jxlD8fD5etLHNvhiki3dIjcMebrz70z4gB/KcvTw+6j1ovOVkUDDYNAuKe/j1JuQLuqkA/CqmERigwN/urKbxy3TDHNGPnvWVkakB/kr3F1qPaNxFdXE8k6RWLJGzeXe4yB0Gw8KtfiXW7TStSitLuM91ODl8/ROcdPKqx/ki2T2Wua2JJ++a402GdML7qq0jDA8tx0qTdsOTqtuh3AiJ/1qrB+pjPsg3b3oq6Vor3lvcXC2l5IsbNEAzqTnYZI2OK5evrfkunjWVgiY5TIMEj7QK7H0+2h4x4J1DhW93mWPNu58/wBHHwI+2uVOLtGubDUZ7S8iMc9s5SUY3G/X4GqmmEHQEW2uOUFS8n7J5qSmklhbDRI/7UeKd/k9cfmLhCfAH3TXskWpWyA80pXwz761ma2NI3SRAxsYzkgZViMb0pEtlJdiB4pUYvylu8yBWxu5e9VZobaTJAyY+U/ZWCW0NzytZFXVj76SnqPHegAjf6Nbi8dDeMzKF94KuDtSC6QiENHduCDkERHY/Kit3plq1yR+U5Gk5VJbIPVRt8qT/JeD+b1AHHiUP4UEt8jRre8Qc7au4VevN3i7fVSF3ps9w/NPqUUmBgF3Y4HzFFJbKdbdxLe2+HHJlucdfl6V53FyCEe4tTjw3/hQF8AVtIA3/KVnj9s/wrX8mRjrqdp/pMfwo73GM881p/o/7qSNvEcn2m2x/imz91AbAcabDn/lSD5K5/Cs/J0Gd78E+kEh/Ci4hgQ5NwmfIRNS0b2UTtHJI7l1KjlhIwSOvWgNwK2nQMcreN8oHrUaam/87f8A7F6PKIFUKJ5eUbAdyo++th3Jffv23x/RpQS5gAaYh29sb/sm/jXj6ekMgPeSyAb7RH+NHXnWOXlWwnkwevuAfdSxvLlwf7nSLt/hAPwoKTsc8HcSX+i6lDNbd8igjmVlYKw9T4HyYV1R2a8aWnEulxQ3UgeUnu+8bZg/+DceDY6HowrlXSLDUr+9SJIJh3n0EV+d2/ZAH29K6K7FeAr7Skmv9QYc84iXkG6qsbcyrn9Js9WHwqoNpinyixL/AEwMp5kXlIwds5FNeHpLnhidhaq8umy7PaFsKjfrIT9HPiOlSGXTL64U80jKh8tqRTQbYt+elZjnfYVtaMqZCuOtX1W8l/uPp1+7Hw5Qqj/OJwar660/tJuZW7qazsQ2BvKC32A10fYaRpMdu3etzL5saF6qNIhY+zxhjnyFNyFqc7S9m3F+ptzajxMSR07tGbHoM42pS27FlcD23WdTuG/qhUH41eTTrj3Iwo9Bih9zNgk5AFJcguCr7Lsf4ficK9vdzN+tLcH8MVKdJ7K+G4QHGn2EbDxKc5+s0cN6sZyDk+VeNqzMOXoPjV6odhDTuDNEgTlUqq+UcaqKLrpujWFuwS0jcsOrbkVFxrDxj6ZNI3GtMwIJpKIh1r2oAKY4cRp5LtUK1KfLncn50+1C6Mh+lQa494kmq9AJCXbbzopp+qT6bayrbyBDMnKzDrigzH3qSuJSVxmpCz25cu5JYknrVT9pXGMguJNH0iXHL7s8ynfP6oP41J+0HX5NK0wW1plr6692IDqo8W/D51BtF4Xjblm1ORwZFY8pXJyMbfPNZznxRcF9kPt9LvLxiYopJT1OATSlxoGrQR8z2M3KNjhScVar3dhpad0ksNigBUKqh5CCPGkJ+JbS2g7zvJZI9mZZIgeY9D16AbYrE1tlOurKSrqVI6ginOk311pt9FeWkpjljOVIP2VOtZXQ9ch73uPZJ3yVkA2PxqC6hZy2Nw0E2MjoR0IoKTtHRHBOvw8R6Il7FgTIOWZP1Wo+pJ2NUF2Q8Qvo3FEUUr/zW7IikB6DPQ/XXQ4j32AIPQ1vB2jGapjcCvGH108MOx6dKR7o5pmdiIXpSgWlljzSixU0rCxALtSiplaWEYyNqUCADpQkA2CEHFKd3SypW4Qmq1AbpHv0pRU9KXWM+VKrESaeoCUUfl1pbuWCDPiaXSDfO9OViyAPCmAySEk7ilhHtT1YDnpW3s+N6BjNYyfClBF4kU6SLfcVuQB5UAIJGOp2pyiiK0llY45gUBPgP0vs++kwpLAKMknYUE7Vbm4suDriK1CtI3LAQWK83MwDYx5gmplwhopvtR1w6nrE4t5HaJWxksMAdNtqrqWN8kqhUMOg3OKsbhzQouJOLF08yvFZojvLLy7RxRjAHz+81d2n8K6RHBpemWEcFva3yHeOEcxAOCWz9Jj6+dcrjs7NLOQ5YmdTFjfwz4fCrP8A5NWpNp/GpsZGIS8t2TrsWUgj8aX7buCbbQNbum09eSCOdkZT4HqCPLNRHs4cw8Xac/MQRdR4I/aFSuJFcUdxcPMGt1y2VAyKiPGl0765OFnZVUqoAHkKlvDmBaZ8OmaEX0NrJdSO6qxLE5IreZnDkh6FiN2lPqBWVLVSADCxLj0FZWJtSGfYfo8umaDPf3Wnw200jrbxBDn80igk/wCkT9VDe2CLOpwP1Bi/2qPcTavLpnDenaXDJi5ktg0jggcoIyfmelRriGU6jwvp1wcmSDmgcnrtggn5Vvihqjnk7ZHuEtQg0nXIrq6d1h5WVyq56gfjS/aR2fcPdoEbanpGpQwatycvMCCJh4K6nf50FukGds0xZmibmQlWHiNq0krJUqKk4m7NeLdDnZZbB2C/qbq3wB3+01FZYb6xkIuLe5tJBsdig+2ukbPiDV4F5EvpJE8Ulw6n66fG+sb6HGp6JZzkjd4wUJ+XSsniLWSzl03TSDDMjY/wsIbPz2++vOWF920+GTzaKRk/jV/XWkdmeoSsJI4YXBIPLJGcfUQaD3fZfwbesX03iJIWY7B5MY/0gR9tT42XuimZLawk6w3sZ8l5HA+vFItZWibi/mX0NqRj6mq2bzscv1J9g1aC4THVWRz/AKrE/ZQi47KuKoAzJGkmPAh1z9aio1ZWyIAlvB3TINXjy+McyyDGDn1rFhZcf3ci/wBOT+FTmLs94jYYl0iTIP0hJFt/rU7h7NNcwpNm/wAPc/A0UwbRX4Vun5aj+qQ/hSkQIcA6kHPgTbuw/CrHt+zLiM/Q08kepH4A0/i7LuJ2IC2Ui58RzH7ko1Yvx9lahOZCBcNn9ZLRs/a1Kw2XOELSXchU5Ui2VfTxNW3adkesMmHt74nx3cD7VFFrLsivgoV9MdwD1luSn+0KejFcSmorEkHmjv8Al8z3a/hS0NosZLFZAB4vdeH1Cr1g7H3ccxsLCI+PeXIb/bNEbXsegAHM+jxN5qoJHzwaejJuJQMUUU7ckAWU/qxB5j/q/jR/h/gzUdYvktbexEbMfeLRc7geiZwPi7D4Gr+03sssInBu755V8ViyoPptipjpmjabpFt3FqiQRLvhFAz6nxNPxiteiG8BdnGm6BF308KvcybMGPM2PJn2yP6oAX0NWjptpEIu8cA4GAo6AUAu9WtIByqhfHmaGz8W3MQ7u1WOIHqSMmqUAslOrah3Ycs3LGm23jUdn1oFyIzsai+o6vfXjEyzls01WdwMljn41oo8E2Si51RyoRX28d6ZyXbN60FEkh35jWxm5R7zgfE4pqIBJ7uQ9WwDTK5kHUuWodcalbpkCXmPkKapc3d4/JaWdxKT0IU4opIKHcs2cgZppJMymiFpwtxHdkGVI7VD+sd/sovZ8AxsAb2/llPiIxgU9hUyGyXeG3f8aTF0ZHwgZ/RRmrb0XgnQoME2KSkeMmSTUgTTtOsYuZLSCMD9VAKTl9BTooj2W+l95bOcL5lSBTee2njB7xeX0PhVjcWX3eymOMYA6eVQq/VizE07EwCy4zkb0wuuVA8rkLGilmPkAMmi1wpB2HWoJ2q6mbPSI9NhblmvSQSOoQdT+FKT1QK26Ia99+Udam1m5RyS/d2SZwOUZC7+Z3J+J8qU1TVhad7bJc80p2mmX3kQ46Dfw23oMt4qKzr7rhTHCxOy4+kfuFbcI8Patxdr8OmaRae1SyNkjPKu3V3Pgg3ya5DopA+OPUL4rPaW1xcMSQZWOFJxggHxry3nWxlMGrabLPCSvOI5eRsfLIroLUeEuGuz3hiXW76wTirUYnUO0uVtYsnGUjGMgeZrbguWw4smki1PhjhiO0aFm7iKzAlx4e8u6/M01FsLRz9qY02GWOTh3UZmt5N2hnTDRHx6bEeoo7xTpFtqeiW+o2jpzcg7xQwOG819CfCpZ2u9kcem2cuvcLd61sm81mfeeIeJU9WUfXVSWckxZeVnZoQeZMFgF89ulDVFIGANBcEbo6tj4EGuqOBNQ/LPCGnai55pHhVZP2xsftBrmfX4FWSO5iPNHMoyR0JFXb/Jxv8A2vhe9098k2swdAT0Vh/u+2rxvkiatFjog5twDtSZiGMgGiMcIJ6eFJmLAxXRRiMlj33WlAhzTtYTjOK2WFqEqAbLGOtb8g8qcCA5yaUEW24pgNVT0pRIiT0p0qADpWw5R5UAJpBnGBS0dvjwrBLg7Y+utu+Kjcj66BWhwkIHUD5mlgsQGTIooeJXZvdUknwG9O4bK7nGeTkHm21BSF+eFR/SfZWsk8QQ4OTS0WlQrvNcE+iCsZNMgyWBf9o5oAZifmOFVmPkK3KTMuTEE/aNLS6tFGvJbwqPXamMl7LIcs1AD6xUxyGaRgRGNvj0FV52w6pEk2mWTHLJHNeOrZOSowmB0+kam9zedxbKhYDk99/j4D5AY+uqQ481ZNQ4plZpCI4rJFOB7uSxY/dWWWVKio8kr7LdOt4+GtQvwH7vUZTbRs594RRjD/XJn6qsfRrgJpGg3EhHPBfm2YlgCebG/wBYB+dVfp13e6Nw9w5GvK1mLGMzRAHnaSZy5YHpn3hVn8O2uj299bSa1fgRWDtdw2vdnNzJgcoHgcY6eZFQugGXbBwvNrUnEMkUacqKGJZwvKQnMevjXOPA8HLxTpqsCc3Uf7wroXjrV5DpV3E0we+1CQyTIhOIU6hSfPA6VSnAdm0vH1rFj+ju2Y+gGT/Cs5dplro630qUx6Z7u5/3VEri+uWfPe26ee+TT+fVFtNHeQnlBHKo9TUMeeHmIcyNg9BWsqM4uiQtfuB716vyrKjvtdso3tpGHx8ayppFbAztY4wutH0zQr5RE9uyPbSqV251IIORvnBNKdn/ABvpHE9ldaQjGG8aPv44WIIZk+lgjr7poX2gaWl/2e6tFco8q2fLfR8pwwK5Vsf5rZ+VVB2VXY0rtB0+4TZo7kJyA4BB91gfkamGSSpDyRpl8XUGCTihF2hydsVJr6EglRvg4oXLaM7HAzjc+Q9TXWc1gLlwT4UrfS8umzAb/mm+6t7uaxhGJrmOLmOFLsBn4UF1vVLCGK4tvb7ZZkibKGUcw2PhUyaSKinZUUs88M7RxSFFAAAAHkK3j1C8Ax3oI9UU/hSV4uLp/iPuFaKtcrbZ2UghFqNxn6MBH+JUfcKOR3XEdlaWc6Q3VvBef8VaJpIxLg4PJykZ32oNYaTqNzpd7q8NsXsLAp7TLkAR8xIXO+dyMV0PoltBY69wvb3Vtn+xvhKTUGDrskzjrj4ikFIqOa+41sLyG3u04kinmGYYpLu7DPj9VS2/ypeHjviiPmi9v1AMgyw9odiMdc82cdKn/wDZrq0XZTwvxdrVy19qKcV95aTP9JIURu9TzwfeGPQU+7cNQ0LgqPVNF4WHLqXEsnt2oXBXDQ2z7rEPLmOT8M07YUV7b9pHECKOa+upB4czoR+7TyPtO1vYGR8eRVTn5DFQFNjgDFe+XxqtwpF+dnnFNzrdm4vCneJyupReUFGGRkedTGO52xnHwqoeyaQrA++/s8H7gqxEuW5P0R8TXRFWjlnUXQdNyvL1FJtdr4HFB+/Zv0wPgK1a4HQ5NOiXIIvfSK/uyOB6MaVF/Kye/Mx+JoMZT4YFavKcdaKFsPry6UoPeyTTBmLHOa0yW3Nehc0x7GKGY4pcRBRzGtUIQZO1eRl7uTCkJGDvzHFAlI1kld2EUCFmPTHU0d0fgW81AibUZmgjO/IPpGi+gz8M6Pbq7u01yR7znGB8KfS8baQg90lviwqXItUKabwboliQyWglcfpSnmNGEto415Y41QDoFUCotc9o2kxD+lto/wBuZf40Gvu1vQYGIbVtKT0Nwv8AGs7Y7rosIxY2xtXscYzg4qpLrtr4bjzz69ZH0i977gaYjt34WWQB9c5MnqYXH+xRaK5L7iRY4h4VG+J9THKYkfpQHROPrXWtPE1jfW17CekkLA49DjofjTO/uTM5PrVxQrBt6xeQknOaGXcXMhNErgEkU1uFPJWlIzkR2dPfCgb58aoHtE1V9R4rvp1fMNt+YiA8hscfOr24yvRo/Duo6nn3oYGKZ/WIwv2kVzRcczpEHJ5nfmY+Jya5szo0xL2xae1uBa20JVw05BjCAbjpn8PjV36Lf2vZLZabw/JZKNb1SNbjVJm35Ix9GBT5DOW9ceVRXsU0Ztc49j1CePmsdDhEzBtwzg+4vzdv9Spl/KE0S51PRLLV7aH2m9tpirvgluVx9Lbr72DUR55NZMmuv+y8Q8KXkCFTFd2zAA9ASNvtqD/yeYVj0bU7lgeaBCjZ8CWA8abdkeq6hPpN9p17FJFdacnO8Un0gOuKV03V7fg/gXWLxVIm1W/ZoVfYIgYtk/53KAKrb2SSe/4lbUeNrXhyxuQGQ8lwBFzAM3r6Cqc7V+H4+DOO11OzjZtOupG5lxtno6/DxxUl/k3rc33GVzquqEmWF3nmJ33G5qQ9r9lBq3Bd13h/nCt38IHXmG5H1ZoatCumULqkUBsby3hZpEhl72FmUqeQ+h+NTz+TDKRr2pw5PK9sDj1DVF/bF1uDSomixcdw1lK235zGShPwBH+jR7+TfeW9pxtNaTvyy3EJiiGNiwOfuFTD9i5fqdFRxjm+VYsANKJhCQWy2OgHStOWeTaMD1rrMDHWOMYByfSvBnbAz8KUjhRN5HyT4DrTyC1uZv6G35F/Wc4pAMOQ9SK1bONl+2jKaSzN+cuOc+USk0/g0WJV5jDjzaV6LAiix3Dn3Ez8KXTTLt929wetSd1srcEPcxgjwRaaS39guQqvIf6zUtgB0OlpHvNcdPAD8acLbWiDn9w/tb1pNfRkgLEq0g1zvkhR8qaYcD0XVrEvuEsfJUpGXUZskRRt8Wpo9wu24xmtHu0/wijHkaYG8011L9IgfOmzxMxw0prGugQThj8qT74k5xgetA7FBCg/TNLQxxKedt1Xc+tMu+33basackBVO3w60Et8Anjy99k0G7n5h3ko5FyfE/7s1Qt3d+1a5qbSMSVgIAycKFTFWh2t3TCwtRzlV71ifeIz7vpVPaY8janqAWJpe8t5eh3xy/S38q58zNMb4L70v2Z4dN9qAEUEduw2BwRCoB9evSivHN5eaToM9zo80kyrGA8sw5e+DnYhfDHLnPpVa8OWfEWp8Wxrp88i2bR2/NzEcoXuk5uu3UGrLmmunnginit7uJye6DXAYlUyoJx02bp40l0OyNdmMi6nb3MmqSTTzCUv72OUDHUChPC3s8nardSWzfmTJMynGBy4G9F7u9uHt7rT7SEWsoOGRY/6QeYKjxFQ7hq1ujxPcoqkSYdcZxgjAP3Unw0UWnrestdOEidRBHkAEdfM0PW7JXKlgM+NNotNu2YE+7sdhSg0+5TICs2/jVXZFULPKwPuuSTud8VlZ7BcFR+byfH3qygBvcvPf2c1nLevItxE8UiN0KsMH7DVScEaVdHtc0uwureRJWvgJ0KbAg+8R6ZGfnVk6ddCXU4oWuBGkzqjNkDlBYDc46V0Le9mujQcZ2/FFtfz2wg9+e15h3Mh5SAd9188A4JrOHLNJPkgF/D+ecAEnm2UdfhQ3VomWE20HKXO8hPRm8vgKNareQW0080LAyczd0cfRH638K5q7XuNbvVNf9g0m6uIbOzYjvIpGTvH8TkHoOldMpKKsxUbHfFPAHG+p66dUuhb3C95lUScARrnoAceFC+0/hLXJ+JW1C2sXe2KjmKkbY61HbTijiWzfng13VF8w1wzD7TUy434u4msdE0PVbHU3MN1b8syuiMGcAZzkfGsdlI0Sfojt5byyXDSRxO6NghlXIO1I9zKDvE4+KmhDcW3rymSW0sHYnLfzdVz9WMUb4e1pNTkuY/YO7kjt2lTurmVQxXGRjmqaRq3RJuAOLbvhWS8i/JlnqdhfKi3NpdplH5TlTnwIyaN6N2nanb8e33E1/a2t8l/CbW5smJWJoMACNfEAY61W0XGMC7vZXisP1L58fbmnlvxhprD88+oRb9MwuPtWikFk44742i4ml0a2tdFt9K0XSMm20+3bIyzBnJJxljj5b70h2m8Urxnxjd8QJaPZxzRxxxwOwYxqi4AyNvOo/8AlvTJdLe/W4lESyiIiSwiY5IyOnhgGmg4h0oqQktoMfraeB9xodBY6Ar0fSHxpmdf0vbLWH/Zyr9zUva6nZ3DsIYbKQohc8txIMAdTvUhsPF43fhS4t4fYVuFksoG3cqR7vpRG17bIwPzuiyD9i4/jQKXTOGuIHjuNR122sHRBFHHHMGAUDAySMmnkfZ3wjIgKcYQE/tJ/GtE2ZvV8skEPbZpZIEmmXw8+WVTUzHHmhw6y+lX109tcLj+mTCnIBHvdPGqxHZdoLhu74ttskbZKf8AmqVa5wEmr6rJejiHT4jPGnuSIGyAoGc82+atSfZMoxZZ0M8U8QliYMpGetKKPOoVwRw/qfDjGE8QWd7ZfSWIIcxn+qSxwPSp3Zd3dHk5lEg6dMH4ValZg4tCe3WlUGa3mtZIzutaoeUirJE5YpHb3c4FRPtT1C54f4XaaEGOe4kEMbg/QyCSfqFWLo4jm5geUuD0x1Fa8e8J23FvCN3pLnknK95bSfqSqPd+R6H40pdFxXJyhda9q3KeXVb6bm2w05+3FBL3UdcOT3jsPR2P3miDQy6bq89hqluYp4JDG6MN1YeFSnSdJsr4ACLc1ycs34KuuL+9BPO2D6rScE+oXDcqSMcnGwq3L7gGK5RvzQIPQ1C4NNk0Se8JhBNjdhJebqqke63wzTqil0M04c1gwLPKZIkO3M5wPn5V5qHD+vaZc9zdWsy+4XUMR7y+Y8/lUlvOLJbqEROq8mOXl5evpTa+1u7v1tbe4kLrYW/coSdwvln/ANdKknZjDgribUOGdZh1KxdlCn8/Bn3ZV8QR54rrDSNTt9V0m31Gyk7yC4jWRD8fxHSuOrrl5+Zds1eH8nLiEz6Te8PTSe/aMLi3B/wbZDD5HB/zq1xSp0TJXyW4ck71pMnMpOKUDHpnpXucqQdzXV12ZIqT+UFe+y8KWung4e9ugD+ym5+3FUZaASanHCN1TGfqyfuqzf5Rd8J+MLKxBylnaFiufF2yf3RVa2OPaXnyBswz8q4cr/I6YqkWnwJrw4P4Dj1EIrvrGplVVh1jhX3iP892+qrK/LEHEnDkr6ZdCKeSIhGbfupMe6SPQ4NRCy7Nn4s7IrO8S/SJdD0VrxUZSe8lmleQ7jptgVC+ANQfSeI7HTbUXlzbXid3dRuMlSfFR6b/AFVUeBSonnZhoV/p+maxqOoBo7y/tnaNJT77IAQGPnk5NRftQ03X9fk0+z0bR72TT+7Xu5VTKHAxuw265PzFX9ounaRxHwzNbalPJZfky3eOScR8rOMEKUyc9COb12oJw5f2vC+k3Ojogmv5Io0iurp2WIIhLR5QZGQG+fjToXuyLcMaXZ9mvBzxatexPrOoJ+cA/vEJGCx8cY6Z8aBtr9tr8SLEpVCrAKfEEkZ+qmHGHCmrcUajPLq3FdtHHJLzcscUkrSf1mOB8APCpRw7p3DmiWKQSWFrq0ka4DvbCM/6X0qOhM53kNzZ6xc2cWFEV5tnqp5iAftqddiPD11N2hz3QkRfyZKSwG5cnIAGPrqLcXhG491KNUEcUlyMJzHCgkHGTvV0dgViLbUNbuFKNC92oiYHOVC5G/zFEF+RU/1LSt9Pn5S0nu5HTNKqru/cJ+b82OwA+NPGlJUkDypJo1b35wCB4HpXSYji09gtejieQdWC53pWTVEUZW1J/bOB9VCbvUVRCsKDA+Q+ygl3eSyHDzKPTNRJgSO64gmRcLPHCP1UG9B7rWmkzmWWQnxY0NiheU/moppSf1UzRGy4b1q53i02RR5uaVgNfyjK30Y61a7uG2AANSex7P8AW5sGVoYh6rk/bR6x7MC+DdXsr+YB2+qk5Wh6lcGSU7tLy+ma9ALEe/I/wq3ouAuHdPi728dI0Xq0sgQD68Uyu9f7KuHmC3nEWhxP+r3qufsyaWwasri2sbqdgIrORvI70WtOF9bm3WzkUE+IxRq+7cuyjTyyW+oT3bLt/NrNmU/AnAoDq/bnw9qiNb6ZbXyoAebM6QE7egJo2CjL3RHsgfbrqKNv8GDlvqFCbgxBuVPrNQLirtjNlfzWtjwhYOUP9LcXUk/N64GBUTvO2HjqcFbOLT9OB6G205MgfFg1HlHrZcqI7n82jOfJVJpV7W6WImSIxKP0pCFH1mudNZ43491EA3vEWslc4wk7Qr/qYFR72iZ3eW5maWU7kysXJ+Z3peYNEW12s6lZiO1tY7y2llWYsyxyhuUcpG5G1VrwxeC04oZ3UOsgdGU+CspFCVvMS4J6qvj5ZH417Iw9oWaM58SfHPxrKcti4pIvThnWli4Ws4IUEYkhUySj6ZAGOX06GpTxVY2t3wpp+paJbm3k7j2kcoO7oQrnJ+IPyqv+yaOy1hI7C7mAt4XJbmODyH3vvzVj3+tRalfXFvp4Een2VjLFAq7LuBk49cVaJY54P1FbGe1vbzux+ULducY2EijOR8cVV3D9011rV7eRjJd3kz1wWcmjvG92bTg/THVyrAPy423NC+zGwlmsZZ1A/OvgbeAqXyy10SuK4nSMgSFT128aUWedUP0ivXc70QXR7knvF2cL5DpTiz0m7lk5iTH0B8aq0JoErcXjgbsAOmdqypH+QZmJBlYEHcHfPrWUWKinzew5DFwhJ+ur/wCA+0heIOEI7PULf2i7tVEMzcwDOMbMR8Ptrl72p35BynAH177UV4M4rv8AhrWk1Gzjik+ks0Ev0ZVPUH/1saIqmS2dBX35JkJBgTB295CM/aaiNzwJwRcFufRNM945JVWUk/Hak7Ptc4Rny2qcPXdmS2OdMSrnx6bgVKtG1ngXXWWPTNUspZn3WITFX+HKd62tPshKXogV12T8Ez8wXT5YiT1gvOnyOaa652Y6Te8NwaAJr9ILeTvIZWXmdeu2cAEb1bUugWLEhTKD6SZ/CkP7H4P0bidfDcA09IlLZHOt32FopLQcQKPISwYP15r3h3sj1PR9aivPypY3EKZDqOYFlYEHwIro5OG51GY9Q/0lNaS8Paj4S28g9R/EVLxxHvI5O1bsi4rjuJZLa2t54+YlAtwgJBO2xoFP2c8ZQqS+g3Z8+Uq33GuxJNG1FPdNnbyAeWKbyadcoff0oE/1SR9xqfEivIzlTQuHdaHDOu6de6TfRMUjmgD27jmZScgHGM4qIXGkajB/S2NzGOmXhZftIrtVrdhudJnGP67GkJEt85l0+fHwH4ijxIPI0cVmzuVGTGQPUijHBETf2RW8Mgws/NCTzfrqV3+uutpI9HkP57TnbH68SGm503hl3DmxgRgcgmzGQfMYNHiDynH95Bd2lzLbsrfmnKnbyNImWY7EfWo/hXX0/DvCVy7SS21kXY5ZntDk/fSL8H8FuQfYNIPxhYZ+yl4mUsiOSHkmKhSqnfwWptr9jPqPDPD15DgMLZoSQ36hwM/Kr6m7PuCJyGbTdJJ9GK/hTm34H4WFnFZG0tPZ4mLRotyAFJ643FCxsl5Ezmy206eOEd5NIH8hIakHCGvarw1fie1uZJoifzkDuSrfDyPrV2zdm3CMzZWGRP2L7H402k7JuGZCSjaivly3eaWkk+Cdg9wlxnpfEWnqXl5JAeU8x95D5MPxozc25Ayu4PQjcEVBoOzPT7C5S60681WCZD9JXBz8cjepjp0Gp2bqkFypix0nTI+rFbq65IdWawzSWlwsy/SU9KsPhxYtTtBPbyIW6Oh2Kn4VDPbL1SBNZ6ZJ68pWlrfWru0fvLe3s4W80kIofQLgh/8AKQ7ILrV4X4r4diR7+FP57boRzTqB9NR4sPLxqh+Ddci0u87jWTPAIzg4iJYehHUfVXUV1rd7O/PLLFk9Qbgig1/Hpt7IZLzTNHuXzu82GJ+sVn4voryIra97TNFgtuTSrATOB/STSAfZVbcS6vNq+sPqatawSyR93MmRySL610DLpfDPPk6LoA2/wWa1aw4bTcaVoI+Fpk0nikHkRzesUAINq+mWkh6ymcvyj0HhSN8kUEHs1rcRzlsGWVWyWPkPSul400KI5istLT9jT1p1HqFim6xW6+XJYRj8KXhYbo5RMUjDZGJ+FSnsll1PS+N9Pvksb1rZnMMzpA5XkbY5IHng/KukbfWkGOQzjHgiRr9woi3EHPaJb8t04DEkvPgN8gKFhp2G6GPelCQD6Vsk5I+l9tNmCSSM2OUE5wPCm2q3Vvp2l3V7K2EgiaQ59BmuiTpGSfJz/wBql2L7jzV5o8NyP3Ix48qgH7qjVqYJtPhjXCzB5FffdgcEUrNPJeXU9zM+JJ2Z2PqTk0voVobrTtRCjMtmVnXHXlzhh8OlcHbOtKi+uD7nHBujQmNGifTYVdWGQ2B49M705vNRghvIDPJbWzkgLIqIJFUHJ5TgnYZ6VC9F4pstK7NtLv7jJkEJgiiHVmRiMfDaotw7qD6xxhDqGtu7q5jeNIpOVVXvFDIP8wtWhLXJbel+0vPe3bSyzxs5BZhyiON3BwcE5zjc7UjxxNHd3Msj6nc6fGYZ5I54AAp5CqoCW/RwR9VSPjzT7WPSdENpFc6faiMteLZEq/dNnk52PX3gPrp1ZJw9f9n94mp2KN7EuNQyxkcxMAQwbqCNs48qoRyzHxzxRYXBlj1iWQI+3MFIIB+HjVt8KcTWXFGl+1t3VvqQOGhRCFkUDr8aA8f9jp9lGqcJanFe2UnvKk8gVlB8m6H571Mew3guDQ9HuJdento7hgZO7eRSqgfqsPHaoSdlMorXk9s4/u0LAI9yqliemwrofsfsZodAMkNtGjzTPKY4veVcnAA9NhXPenRpqvHl1IOYwSXU0xz+qCzD8B9VX32f9snAXAPCi2M6X2oakqqrxwRYCYHTmbbrTxy5CSb4LQ07TtYkJ5bcYYDeRcAfCnDcMarcnEt3HEM5wP8A0KpzX/5VjZlXROFoY/dAR7ucsc+oXAx86r7Xv5RnaNqRkW31C306JxgLa26qR8zk/bVvISsbOqoOBbdj/OJmlx9Is2FFKTWvZ7w4gk1zXtFtcbhZrxAT8FySfqrhXWePOL9XbOo8Q6lc5GCHnYj6ulAGuZ2PM0rsfU1LmPxHdOpdu/Y9opMdrfXOoMDgCztHC/6TBQaietfysdBtuePQ+ErqdhsrXU6xqfkob765A52Pr55rYMQcjApbF+OJ0JrX8qnj27kYaZZaPpsZGAFgaVh8ywH2VDNZ7a+0rVR/OuLL5FO5WB1iH1KKq5WI61sGOd+lTbHqSW+4l1fVWZ9S1S9uS3XvZmbPyJptFcgYbA2O3KuNqDA4weXNLRyOVA5RRbDUOLeMSQUCgnIC+FLR3pQHHNk4oKpb9XBHlS8I812+NFsepIbPW7yBx3TnboHANSbS+Op1ULcwRShOpACnFQSHkGeYNsNiKcfmiQWZTgY8zikLRE61fX9N1rRruzY900iZTO/vDdd/iBVT3UhEnl1FH2kiyAB0+2gepxKJCVx50EONDbn6HxyRRC0PehV9aEk+fjTuwmKFQPA0FEp4Wvn0rUV7527iT3JQDglfP5dauzRNR0a04Znm9vjkmmhYc+fcOem9UVHHHdQ56P1BFIG2uh7n5wjO4U9aLkhaku4k4gm1u7ttNtpGkggJCYOckmrm4HtrXTdHtIZYlDJGA++SSd6qXgHRVN4t5LCI1UDAG+/rVlG4McJAbB8Tiq1vkV0ibflS3U8nMgUeA6kUnLrkfvKgCA75zioU97Lj6TMehztua19qlD+9l8Ntgb4+NNJCtslcmulBlWOM42asqK80zSFu7fffpWU+BWyo5rmHCsr5wckA5G1IyXMDHYnOMEbbeNCFLdyVKdevmK9ZJWBAGANzTMwjcXK92FLErkYPNg/UKF3CObrvoSQ+crg7g+BHrXhEiYLHmHXBJBFYAS2xY+Iwfo0pcjTokuj9pfFulxokt019GuwMzsH2/rg5PzzUz0ftvl7pBdyXtu3Qho1nQfPY/ZVR3Cqy/pAE5PofOhdzG8blj0J3A8KjeSNEkzpXTO3CzzyyajYgePeROh+o7Ubi7Y7G4A7q5sWXzWdR9hrj64DoxKsQOtNyzY6mqWR+x+NHaC9qNs7bRrKPNGVvurY9pdmPpREfFDXF0c80Ywkjr8DinUWsapEMR31wo8hIae4eI7FPadpoOGjP1mvR2naOfpK4+ea5Bj4l1tF5RqEpH9bf76VTinVgMPOsn7SCq8iJ8R16naLw9KPeLY9YwaUHHHCb7M8I+MIH4VyCnFOohslYSPLlI/Gl/wCyy6/Stojn1YU1kE8TOvE4l4OuDjvrLJ8xil47vhSb6DWn/a4rj5OK5f07cj4SGlo+Lyp2il/7QGn5BeFnYa2egTkd2sRz0Al/316dF0xs8sbj9mSuSYOOTGBlrlceWD+NFLPtKaPYX16nqM7fbQsl9k+JnT0vD1iVyrXCn4g/hQ270aBAe7uXBHg0QP3GqRsO1towAdcuxj9ZGP4Ubs+1izmIMus25P8A0qEfeBVKaF42WE+j3OcpLDIPIgqaQnsbi3wZoSB4HG1R+17UNM5Qp1bSc+ZmAohF2k6a4HJqujn0E6fxqrQtWPeVdugx6VhHkdq8Tju3kwfbNKIPT84h/Glv7MIWGQ+lt/np/GnsiaY3Y4G1Jlzy9c/KnTcWQnqunfWv8aTPFNv+rpw+LKPxo2QUxoQfHNaM2P8A3p43E1p0dtMHxlT+NIycSaXnefSwf8an8aNkOmN+8ycdK3VyvStW4n0VN3utJHxnQfjWh4v4fX/8R0df/wCwv8aVoVMeRSsN6fRSkjYH6qCDjzhtNjrOjf8AeE/jTTiLtT0jTdMe4025sdQuchVhilXx8dqTkkUoNsl8TTMNopD8FNVr23cQvbacnD8ZKTXGHnHQrGOg+ZqI6v2ucb3IDRXkNhE/RYYhkfNqhN5fXV/eveXtxJcTSnLSOeYn51jkzWqRrHHTti8O8i5Gc9aJcKTywanNbRzJA10O6y68wOdsH7OlC4z0YeFJXUjx3Mcq7EnIx4EVzmrLN4N4a0zivh38mapM9qNOvpCDvkLJhsEY88/VU70Pg3s94cK3MVpeavdocp3vuRK3mBnr8qrvs41aK24jEYkZ7bVIxGpf9CUAFR69WXPwqSdoXEn5B0sJAVOoXGY7dD1XwLH4Z+utINJcky5J7acRHjLiDXLKRo4Y4tO7iSOEkIh2wB54AFRTsy1h9G4ku9DvLgXqMCWYvzCaFvdYHzIr3sh0+XTezbVtXmYm51C4jgDnZjvzMajnYlp5uu0rVLyaId3bG4HmB7+5qrGkgrDqCcPcQapwnJcNPpaTc9s3UrG264+HTFLcSyrYaJd3JK5EZ5SB1J6GpDw5wfHxNxFJM9jDcKNndiwK7nHSon/KXGl6BPa8J6VM7tEoe8y2eRj0TPp1+dDbSCuSruFiI728nOSGTkZSeqk8zkn9lKhlw7PM7sxJJJyanNuRp/B19dOuJLwCFD4jJG/1KfrqGBF5sb1mjQbKCa25D5U9WHyTbzpbuFGKYA1YmzvW3dt5UUFucZCE/LalBAp/RPyoAFLAfWt1tyRRM223ug0stnJ1K7eeKABi2+wyKUFso36ii40+XoygYNKxWSAYLkrnfHSgVgYQr1AHptSoiwM8oo/FZwcuFRQc9WYCl1tVA96OPDeWTQJsjSqSd87elKckoAxGcDr6VKo7BWHuRbH0pzbaZKJO8WAEfpEjfFAWRHkmc4SORvOsMU6sqrEc46Eb1NotKu5GKrFgnx8T9VPbfhPUZmDGFgxb9XFAWV1LFdkAd0wPwxTaS2uyvvRHOPGrig4AvpZAUQt6g7UZsuzUMqi5nQnp7oO3xoA579hvGORCTnpSkWnX30ltXb4CulIOzi0QK+EZRsR45o1Z8D6YmFa1VMDPxpWFHNGn2OqrjlspiDvsKmfDXDup30qd5bGFM7l+tXsOFtOjUJHFFGUPViNxTyHSNPi5ELRc3X3d/u2pqSQtWRPStCSztBHBDkLjm5jsT5mlZNP51DSJIY+hCLuBU0s7SIRAlhybcwRPeJ+FLmOLn7pDjfoBuvxp7i0IOuiN7rwxyuGc8uVP8KexaEJJti6Z+ioUD7ak7rFBnZy2fdCkZPwxSqmLn5liYtnAww/9Cp2K1RH7XQbMLiWRySM7jesqQTvCj8r82fRSftFZRYUjj0xhlJbIHj61tG/nsT0UnNIyPkjGCwGDScnI2CxJb+qK2OYXnuVZTmLG4OdsmkmlZD9HYjGdq0LADuy/jjPp9dIsASFBGB0yDQFGzMwYyHJx4U1mbIyOh+ylXcgjmJAA2z40hIAVYlseWPOplGy1wNJ0Qrt40PeNlJx0p9OjnyG2+KbMj5ON6nRmkWNyNq0NKlTnpWrK3iMUqNLRpWVty7ZxWYwKKAwCvSRWu9eYNMD0mvMmsrKVAZWVlZRQGVmTnNZWUAZk+Zr0HFeVlAG3O3mfrrOdv1j9da1lAUj0sx8T9dZzHzP115WUCpG3eP8ArH663hlIb384pKsoCkO1lib6TY+IpRRbt/fQKYVlAtQ1BDbPjEkef2qJ28CYwHT5Goln1rZXYfpH66VCUCSa8Fjjt1Hhn7qHI4xjJzTKOVmBVmJ8s70orMCD5VLQ2qCsDgoN+vjSkid7Fheq7/OmEMuPh1FOo5uXcnagmwjw/O6ym0Z+6L4MUmf6OQbjH2j6qtDT9G0/tBuItUmuhBrNughvY3+jzDpJjyPifOqjeATRAhsODkY86P6Jqd+l3DdWE7WutQ56NgTLjfAOxJ8VPWmmN8nRnEUEOk6Fo2gQLymCM3E4Hg7bD7Bn50N4D0aXS7TXr1YGEtwr8gA3bnfbH1E1D+HO1eVj7NrOiGS6TAZ7ZwGJ9UfcfDNSK+7TLqWwS2tNGFlEp52uLyQJj4AE83n1A8atNE0yYcEahecLC+mIiWe4h5YQ7D82R+m3oPKua+0i8TVuLZO6lednkJkdjkvk7n571KePOOTqFq1palZJCeQyc2Cx+Pl6VD9Gs2hYX9wpaTBIUjJc0Sd8FJDXi+VZPY9KgI5YV5mAH6ZH4Cgq2BXYoBmpBHpl1cTyvIORmcvkrsD9dPrbSpGUKFDgjr40kWRuOwOF5VJGN8Clo9PPJk83rtUyg4alcoI7WRBsCxB6efwovFwhO3vYVVA89/iKAK8S0blwQQc4G1LR2aKDzAsfAVadnwLDM35z38DdUzg0Us+BtNhkSNyvec2ydD0zk7GgVlPjT5ZPeS3ZFAwT5etOLfSnf6XOcgAeZq8Twva8xK+xoikAnGW+PhT2DhyyT6Eao2z+6nXB3ORnOceYoEUnBw/dzHCQsc5HvjYUTs+ELtolYrkdMAZHrvV3ppVmQri35snlz3fuAHzB6/Glbu0gtvzcfcqmQBEu/wA/LwqbCioNN4FuZHZWiI5SDgqdgfhR627PXU8zg+8dlC9Ksi3mBhZBGEbIyyjdsdPu+2nQ71piWR+T9Qj7tqLCiA2PAMPeyCZlHL0HMcketGbfguyiGGPMq4fx+qpaGCTmRbaMyE4ZncKR57VrKqiVQ8/JzjDBRgfWaVhQFh0W0hKdzEkavg8xXalRZQJJy92zN4jw9MUQmihkbkQysYz+k3X126is5WUGGKZnRdgQSAN9wTQUI28Tq2RiJQcFQMEVvNzs4jMuCGIJxmljC00DmJokAbZj7x9fLP10nERHMwXv5G6EhAN/kDmgDQQTq3vSSShzgc+4UfCtXt2ZkUyMzZ2Zjt8K2jivpVLOMjJDsX5ceWRjyxTiOzcqYZrpzjqo2yPjmgBBA8UeeQ58s4Ary6N7JKO7RCAN1IyPr86fR28LTlouaMAY3O7H4AVpdCKCMFkZ2Y+8GYgH1oAZxveA/wA6McW3Rj09aWNzHGeVe75Sv0ubH++tZLi2mAdZYw2ACAu/lvvWPbgszsFwmACd8qfQUAbRyM7Kixp3fXY7E/OskukgVTIrc+SF5B1pPu7yN2Au1VDj+95A9Rv1pUA8w55ObGx3yc/AdPDzoAQkKOeZmkYHP0V8aynZtBLszOvj9LcVlFio44j3fvGXlDb7ikmTLknlIxuPup/7KScEo+BkkjpnwrGtlQ56tyjAxnYeVdBygwIcczbnmIwOteYJyRvjoG6UQ7nmBJXkyTjGxrx4m5AFUOQu+M4HxoKBso5cAK2SAOm2aSkG4LHA67jFEGTuyYwFLEgE4zg0jNEQzHc58cbUDGIw2VIw310m0XLuBj1p3KgA3Vg3X1NJiJskyA4PvDagBnyqSeZQPWm7jHhnfyohLDleZlJx1AFa9yOQh1IAO2DnNIdg5k2AI+oVqY1xjBJ8sUQ9nC8xwD6A9K8MAyBuD4bb5oovyA7uxjJOK8MYx1p/LBkHqo8c7b1kdu3eYIIGOuNqKDcHd0fHpXohOM+HnRDulBYMenn1xWxh5RkDCnf5UUG7BhjPLnFeGJgBkdaJiLJz089q8KENynr5hQaKDdgwxsDgjFZ3Z3oi4B5Q25+FYyLkZGB06Ypai8gNKMPCsCE+FE5kG+wCgDYUl3YIXCEknrnpS1DyDAqRWcreVE3i5se6MDy8a0WJSPeGdtiKNStweFNeEYoo0EfJkDfp86SNuDknOB4UahuhhWU7FtlcgivPZwTgH50aj3Q1rKcm3I8a8MGwxvvRqPYb1lOGgdRnlOPOtRCSM0ahaEkyDmnEbZWtRC+dq9ERU7ZzScRdiqtty+PnTiJwVCt1pnk5xjevVkwxwDms2JoL2s3dkDwPWn6iOTlZiRjo3iDUfScZAJ3p7aXJjbxPy60AHo9T1GIYZ4roDYGVcsP87rWqX+pNF3LFSWYnmJJ6nPQ+VNrZp7hhHDY3EzE9EjNTDh/gjifUJA35OWyiB98zt72MeA+VAdgPSNNlnulB5pZ3bKqAdzVoafwFP7PHNclZse8AhIx4ePlUs4S4DtNAIu2uUnPIO95lB5lI3xj6qm0Ua28ThEQu3u4cYBB8s5ospIr+24L09bRUlgIGB9FS3eg/1vCl7bhzS7dGEb+y8i5kUOCACN+u5x9dTiWKIQRszJ+cPKUXx2wT6gUOu5AiwSy8k0RfkkKxlubf3d/LzosYwttAtI7fvF5e7Tbm358eG3jtTqDSoIp1SOJCEXKsRjqdhv4+NPVeW27xUnMYU5IzzNk5GNvD09aXim5RyusaA5IVz77eAwPE/GiwGstp/NXmhlVixOHCglh6V7ZckHMiRhkkzzSyjG+PDNOCbgukKoDyg+6o5VONhtW7+1pM3euR15uX6JG2wzuCKdsKPIopk50xbyMXycIBt8T1pMW9wZi0byBOXJRdlwBjBxSkq3E0YSF2BJ2JfGM+o3rdLa6UGIuZFQYK+83MPDc0goRksYpHZXhGCAQjE8u3300Szjhkw9qUUgb4CrnfGPEU/fvUtwjQuOQ82cbmvEu3nAjiyoH0gThh8/GgD2GeGCNo4YXRjuDtg7+fjisN3LGSC6bkA4XPXxydh86WdxCBCBzEjBYKTvv49fOmt/OpXEjIF6qpXmznrnyoA9uJXTGIY3lPjJMC32bU0uGuxGxfO67KH8fIbfbSMcroWiUJKrDJVVGF8t9sfKnVk8qFu/nWN3bmCq26/V4UAa2000aB5S0YAIXmbnbenCzQxSsomEjt7xAI6/AVrGvePJIZw4K52Ykt6da1BijunkXulEileZAE5d/voAVa6bmHOVWHYNhTnPjnalDegovs08bKDy+6d/himsssCRNm4EzlsDnUjf49R8a3W6iiwZY5OckZIYkA/Hr86AHDXD93yLE0hBOS64GPwrIriZMNiIvsQANwPj0pKKWOVpD7WwYNzElixwfDB8K9nuUjdkgV5W5egUDIPlmgBZtRKO0a5Dn9boKQmmZpgzhXK5ycDp5ZNaiL88ylkEi45ecgsMjO2Rt1rWSGSVSJeYADJkACkbjO49KANoZ4FTnA5+c77fbSnPAx5UBOfd2x4DxpvPDYSlUQs7DYnm5t/Umse2SKRFYKkTEu6q24NADn2iAO0MbySPyg4GME0o7IikGRImPUA9TmmcVswuGKmPu13UDrn1PWtXS4MSxi3JAByRhQd9xnrQA8llMRIW4SNQdgymspq8gHIMcuFAxgHHpnc1lAHLrQ83K0UrRr9JjyZJ+30pSOXmjDKC/KcHKkHIHQbbZrKyug5DSePmUztGFBxzgjcf8Aomklj59+bnUjfcLn51lZQOJoYDLHy2+Hw2cjxPj9VetZS/QMUmBkHC5BH61ZWUFCcloHYhkk+WR+FeGwjEmYyCmT3uWLY+wVlZQBp7IwSRlZByDA94A48jmk57CcAOwBVtjyjJ+ysrKAEfZXZ2LBFVV5clft+yvRar3zNzYXooA9KysoA0aBgCWiyhwMEe9SfLH3hCsxXHUDI+FZWUAamEGRUfkRsZ6gZFaOqNgh8Z26/wAKysoA1cHmABBB2yRuPSk4YUZnfmZcb7g7eVZWUAZLCNvpAZwcHY+te+zIygopxuWJ3NZWUAbGzZlDAD3snO1ePaMFAwC56sRsBWVlACaQELzKVIzsCa1eGZlHMMAHw8D+NZWUCs99mJYsxYqN9xgGtlTMeAmDjasrKpDsSeHlPOeUDwwRvWwgYDPKDnyrKypGjSWDMZdQVXx9a8mjcZyAMdMb5rKymJmywSMjcsUrMfEKTS9lo9/ehFtbSWUscDCk5rKykCDFvwNxLPEeTSpTnYgruvyolD2W8UtIkT2+Gk+iSrN9oGPtrKyobo0iFD2M61BHzXc9vCWzyLgktgZPSiMfY3Ekcb3eoMpP01EZyNvD/fWVlRZQU0zsu4fWCOQTrc4kCsynIOSR1UnFTXTOD9A0mzZX063W5fGMRhgceXruKyspMoN6dbWdkO7iskaUqq47sh2Gd/DG3oaJTXE0cZMFqokSM/nDEeZAdhkY69fGsrKQCsKrdxrdTXE0xeMKWjARB/VwNx8zSU8U8kMrAM0sgBQTy8gU7EnAG43+ysrKAMtIZ5mjuXlRQqMr7AYY9cHOSK3ijhLOjN3yc3uFgSPq6AelZWUAO5LNkcusTSDOWEakkjz8qZT2tvlQiEqTykJGMgncHJzWVlADuNpHjcqJI+XbMhGST6/wryO4HPGpXmZyBGRhgMDy/jWVlADqWW4hj5onh53wCj4UD4gZpsj3YQzXUsxwCSse4UY8M461lZQAqLy2BDGGViEJJky3rj06ZpGO5E0avFFFCATyl5s83XI2FZWUAJTqZWa3Mp8Cnkc+R8xitUtJTBK07gsFLL4nbb0rKygD1rETwustysYdcCNVH18u9bS6THNGicgZSgUuMsSN8DHgKysoA3tNOjgy8ZwCSDygEj7sUlNAYz3E0g7xh15Bv5Y8jWVlACUmRIGQStKqgAkZ8d84608xKISVtzyMu5cAAHzGTkVlZQA2J5pmZZUUYA5QBzDpkZG2DS0PtDsEjEQUNgFveKjzO1ZWUAMr6Tu5B+ZllkDgcwVV388ZpZJZSHUxOAD7oZz4fdWVlACkkqgyBSFCKA2JBzUkwxHHHaxKnuljIzfUc/76ysoAbzXRiiQz3JuFfKuw9zlPjnqaSs9Qtxbl5buBmUkKgkJ5Rnoc/wABWVlADq31SJ5XhtzDchMHlBAGD49KysrKAP/Z"
st.markdown(f"""
<div style="background:#0a0a2e;padding:1.8rem 2.5rem;border-radius:16px;margin-bottom:2rem;
     display:flex;align-items:center;gap:2rem;border:1px solid #1a1aff33;border-top:2px solid #1a1aff;">
  <img src="data:image/png;base64,{CAR_IMG}" style="height:70px;border-radius:12px;object-fit:cover;flex-shrink:0;">
  <div>
    <div style="font-size:1.7rem;font-weight:600;color:#ffffff;letter-spacing:-0.5px;line-height:1.1;">LL Rent a Car Joy</div>
    <div style="font-size:13px;color:#7a94ff;margin-top:4px;letter-spacing:0.5px;">Dashboard Financiero · Ingresos & Egresos</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    st.markdown("---")
    st.markdown("**📂 Tu planilla**")
    planilla_file = st.file_uploader("Subí tu planilla Excel o CSV", type=["xlsx","xls","csv"], key="planilla")
    st.markdown("---")
    st.markdown("**📅 Período**")
    year_sel = st.selectbox("Año", ["Todos", 2025, 2026, 2027], index=0)
    months_sel = st.multiselect("Meses", list(range(1,13)), default=list(range(1,13)), format_func=lambda m: MONTHS_ES[m])
    st.markdown("---")
    with st.expander("📌 Formato esperado"):
        st.markdown("""
Tu planilla debe tener esta estructura:

**Egresos** (columnas A, B, C):
- `Fecha` | `Concepto` | `Monto`

**Ingresos** (columnas H, I o a la derecha):
- `Fecha` | `Monto`

Ambas secciones en la **misma hoja**.
        """)
    st.caption("LL Rent a Car Joy · v2.0")

# ── DATOS ──────────────────────────────────────────────────────────────────────
using_demo = planilla_file is None
if using_demo:
    ing_df, egr_df = get_demo()
    st.info("📊 Mostrando **datos de ejemplo**. Cargá tu planilla en el panel lateral para ver tus números reales.")
else:
    ing_df, egr_df = parse_planilla(planilla_file)
    if ing_df.empty and egr_df.empty:
        st.warning("No se pudieron leer los datos. Revisá el formato de tu planilla."); st.stop()

# Filtrar por año
if year_sel != "Todos":
    if not ing_df.empty: ing_df = ing_df[ing_df["fecha"].dt.year == year_sel]
    if not egr_df.empty: egr_df = egr_df[egr_df["fecha"].dt.year == year_sel]

if year_sel == "Todos":
    # Agrupar por año+mes
    def agg_yearmonth(df):
        if df.empty: return {}
        d = df.copy()
        d["ym"] = d["fecha"].dt.strftime("%Y-%m")
        return d.groupby("ym")["monto"].sum().to_dict()
    im = agg_yearmonth(ing_df); em = agg_yearmonth(egr_df)
    all_months = sorted(set(list(im.keys()) + list(em.keys())))
    months_labels = all_months
else:
    im = agg_monthly(ing_df); em = agg_monthly(egr_df)
    all_months_raw = sorted(set(list(im.keys()) + list(em.keys())))
    if months_sel: all_months_raw = [m for m in all_months_raw if m in months_sel]
    all_months = all_months_raw
    months_labels = [MONTHS_ES[m] for m in all_months]

if not all_months: st.warning("No hay datos para el período seleccionado."); st.stop()

iv = [im.get(m,0) for m in all_months]
ev = [em.get(m,0) for m in all_months]
nv = [i-e for i,e in zip(iv,ev)]
ti = sum(iv); te = sum(ev); tn = ti-te
mg = round(tn/ti*100,1) if ti else 0
mejor_idx = nv.index(max(nv)) if nv else 0
peor_idx  = nv.index(min(nv)) if nv else 0
mejor_lbl = months_labels[mejor_idx]
peor_lbl  = months_labels[peor_idx]

# ── KPIs ───────────────────────────────────────────────────────────────────────
st.markdown('<p class="section-title">Resumen del período</p>', unsafe_allow_html=True)
k1,k2,k3,k4 = st.columns(4)
with k1:
    st.markdown(f'<div class="kpi-card"><p class="kpi-label">Ingresos totales</p><p class="kpi-value">{fmt(ti)}</p><p class="kpi-delta neu">{len(all_months)} meses analizados</p></div>', unsafe_allow_html=True)
with k2:
    pe = round(te/ti*100) if ti else 0
    st.markdown(f'<div class="kpi-card red"><p class="kpi-label">Egresos totales</p><p class="kpi-value">{fmt(te)}</p><p class="kpi-delta neu">{pe}% de los ingresos</p></div>', unsafe_allow_html=True)
with k3:
    cc="green" if tn>=0 else "red"; ic="▲ Rentable" if tn>=0 else "▼ En pérdida"; dc="pos" if tn>=0 else "neg"
    st.markdown(f'<div class="kpi-card {cc}"><p class="kpi-label">Ganancia neta</p><p class="kpi-value">{fmt(tn)}</p><p class="kpi-delta {dc}">{ic}</p></div>', unsafe_allow_html=True)
with k4:
    mc="green" if mg>=30 else "amber" if mg>0 else "red"
    st.markdown(f'<div class="kpi-card {mc}"><p class="kpi-label">Margen neto</p><p class="kpi-value">{mg}%</p><p class="kpi-delta neu">Mejor: {mejor_lbl} · Peor: {peor_lbl}</p></div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── CHART PRINCIPAL ────────────────────────────────────────────────────────────
st.markdown('<p class="section-title">Ingresos vs Egresos</p>', unsafe_allow_html=True)
st.image(chart_main(months_labels, iv, ev, nv), use_container_width=True)

# ── CHARTS SECUNDARIOS ─────────────────────────────────────────────────────────
st.markdown('<p class="section-title">Análisis detallado</p>', unsafe_allow_html=True)
c1,c2 = st.columns(2)
with c1:
    st.markdown("**Tendencia acumulada**")
    st.image(chart_acum(months_labels, nv), use_container_width=True)
with c2:
    st.markdown("**Margen mensual (%)**")
    st.image(chart_margen(months_labels, iv, nv), use_container_width=True)

# ── CATEGORÍAS EGRESOS ────────────────────────────────────────────────────────
if not egr_df.empty and "concepto" in egr_df.columns:
    st.markdown('<p class="section-title">Egresos por categoría</p>', unsafe_allow_html=True)
    cat_img = chart_categorias(egr_df)
    if cat_img:
        st.image(cat_img, use_container_width=True)

# ── TABLA ──────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<p class="section-title">Detalle mensual</p>', unsafe_allow_html=True)
rows = []
for i,m in enumerate(all_months):
    i_=iv[i]; e_=ev[i]; n_=nv[i]; mg_=round(n_/i_*100,1) if i_ else 0
    est="✅ Excelente" if n_>0 and mg_>=30 else "🟡 Positivo" if n_>0 else "🔴 Negativo"
    rows.append({"Mes":months_labels[i],"Ingresos":fmt(i_),"Egresos":fmt(e_),"Ganancia neta":fmt(n_),"Margen":f"{mg_}%","Estado":est})
rows.append({"Mes":"TOTAL","Ingresos":fmt(ti),"Egresos":fmt(te),"Ganancia neta":fmt(tn),"Margen":f"{mg}%","Estado":"—"})
st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

# ── TABLA CATEGORÍAS ──────────────────────────────────────────────────────────
if not egr_df.empty and "concepto" in egr_df.columns:
    st.markdown('<p class="section-title">Detalle de egresos por categoría</p>', unsafe_allow_html=True)
    cat_rows = egr_df.groupby("concepto")["monto"].agg(["sum","count"]).reset_index()
    cat_rows.columns = ["Categoría","Total","Cantidad"]
    cat_rows = cat_rows.sort_values("Total", ascending=False)
    cat_rows["Total"] = cat_rows["Total"].apply(fmt)
    st.dataframe(cat_rows, use_container_width=True, hide_index=True)

# ── EXPORT ─────────────────────────────────────────────────────────────────────
st.markdown("---")
col_dl, col_info = st.columns([1,3])
with col_dl:
    st.download_button(
        label="⬇️ Descargar Excel",
        data=build_excel(all_months, iv, ev, nv, year_sel),
        file_name=f"ll_rent_a_car_joy_{year_sel}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
with col_info:
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    st.caption(f"Reporte generado el {now} · LL Rent a Car Joy")
